import streamlit as st
import pandas as pd
import re
import unicodedata
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook

# Coluna que se edita na Tab 1 (a que o KPI "Cumprimento do Plano" usa).
# Na VSP existem DUAS colunas com este nome; usa-se sempre a ÚLTIMA (bloco Projeção {ano}).
ALVO_BASE = "numero de acoes de formacao a desenvolver do curso"
COL_EDIT_LABEL = "Número de Ações de Formação a desenvolver do Curso"


# ════════════════════════════════════════════════════════════════════════════
# Helpers
# ════════════════════════════════════════════════════════════════════════════
def _base_name(c):
    """Nome normalizado: sem sufixo .N do pandas, sem \\n, sem acentos, minúsculas."""
    base = re.sub(r'\.\d+$', '', str(c))
    base = re.sub(r'\s+', ' ', base.replace('\n', ' ')).strip().lower()
    return unicodedata.normalize('NFKD', base).encode('ASCII', 'ignore').decode('ASCII')


def _achar_folhas_ano(sheet_names, ano):
    folhas = {}
    for s in sheet_names:
        sl = s.lower()
        if str(ano) in sl:
            if "vsp" in sl:
                folhas["VSP"] = s
            elif "ld" in sl:
                folhas["LD"] = s
    return folhas


def _col_alvo_pandas(df):
    """ÚLTIMA coluna cujo nome bate com o alvo (na VSP há duas -> a do bloco Projeção)."""
    cands = [c for c in df.columns if _base_name(c) == ALVO_BASE]
    return cands[-1] if cands else None


def _idx_col_alvo_openpyxl(ws):
    """Índice (1-based) da ÚLTIMA coluna cujo cabeçalho bate com o alvo."""
    idxs = [
        c for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None and _base_name(ws.cell(1, c).value) == ALVO_BASE
    ]
    return idxs[-1] if idxs else None


def _orig_val(v):
    return None if pd.isna(v) else float(v)


# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — Editor do nº de ações a desenvolver (com filtros, preserva o ficheiro)
# ════════════════════════════════════════════════════════════════════════════
def _editor_pfe():
    ano = datetime.now().year

    pfe_file = st.file_uploader(
        f"Carregue o ficheiro da projeção anual (PFE). "
        f"Só serão editadas as folhas de **{ano}**; as de outros anos ficam intactas.",
        type=["xlsx", "xls"],
        key="upload_pfe_reforecast",
    )
    if pfe_file is None:
        st.info("⬆️ Carregue a projeção anual para editar o nº de ações a desenvolver.")
        return

    file_bytes = pfe_file.getvalue()
    try:
        xl = pd.ExcelFile(BytesIO(file_bytes))
    except Exception as e:
        st.error(f"Erro ao ler o ficheiro: {e}")
        return

    folhas_ano = _achar_folhas_ano(xl.sheet_names, ano)
    if not folhas_ano:
        st.error(f"⚠️ Não encontrei folhas do ano {ano}. Folhas no ficheiro: {xl.sheet_names}")
        return

    st.success(f"✅ Folhas de {ano}: " + " | ".join(f"{t} → '{n}'" for t, n in folhas_ano.items()))
    st.info(
        "✏️ Só a coluna **Nº de Ações a desenvolver** é editável (a que o KPI Cumprimento "
        "do Plano usa). Use os filtros para localizar o que quer alterar; as edições são "
        "guardadas por linha e mantêm-se ao mudar o filtro. As fórmulas e as folhas de "
        "outros anos são preservadas no ficheiro descarregado."
    )

    if st.session_state.get("_pfe_file_name") != pfe_file.name:
        st.session_state.pfe_alvo = {}
        st.session_state._pfe_file_name = pfe_file.name
    if "pfe_alvo" not in st.session_state:
        st.session_state.pfe_alvo = {}

    for tipo, nome_folha in folhas_ano.items():
        st.markdown(f"### 📄 {nome_folha}")
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=nome_folha)

        col_alvo = _col_alvo_pandas(df)
        if col_alvo is None:
            st.warning(f"A folha '{nome_folha}' não tem a coluna '{COL_EDIT_LABEL}'. Foi ignorada.")
            continue

        df["_row"] = df.index + 2  # linha real do Excel (header = linha 1)

        store = st.session_state.pfe_alvo.setdefault(nome_folha, {})
        for r, v in zip(df["_row"], df[col_alvo]):
            store.setdefault(int(r), _orig_val(v))

        contexto = [
            c for c in ["Centro de Formação", "Código curso", "Designação completa"]
            if c in df.columns
        ]
        tem_centro = "Centro de Formação" in df.columns
        tem_cod = "Código curso" in df.columns

        fc1, fc2 = st.columns(2)
        sel_centro, sel_cod = [], []
        with fc1:
            if tem_centro:
                centros = sorted(df["Centro de Formação"].dropna().astype(str).unique())
                sel_centro = st.multiselect(
                    f"🏢 Centro — {tipo}", centros, default=centros, key=f"pfe_fc_{tipo}"
                )
        with fc2:
            if tem_cod:
                codigos = sorted(df["Código curso"].dropna().astype(str).unique())
                sel_cod = st.multiselect(
                    f"📚 Código curso — {tipo}", codigos, default=codigos, key=f"pfe_fcod_{tipo}"
                )

        df_f = df.copy()
        if tem_centro and sel_centro:
            df_f = df_f[df_f["Centro de Formação"].astype(str).isin(sel_centro)]
        if tem_cod and sel_cod:
            df_f = df_f[df_f["Código curso"].astype(str).isin(sel_cod)]

        if df_f.empty:
            st.info("Nenhuma linha para os filtros selecionados nesta folha.")
        else:
            df_f = df_f.copy()
            df_f[COL_EDIT_LABEL] = pd.to_numeric(df_f["_row"].map(store), errors="coerce")

            cols_show = contexto + [COL_EDIT_LABEL]
            col_cfg = {c: st.column_config.TextColumn(c, disabled=True) for c in contexto}
            col_cfg[COL_EDIT_LABEL] = st.column_config.NumberColumn(
                COL_EDIT_LABEL, min_value=0, step=1, format="%d"
            )

            fkey = "c:" + "_".join(sorted(sel_centro)) + "|k:" + "_".join(sorted(sel_cod))
            edit = st.data_editor(
                df_f[cols_show], use_container_width=True, hide_index=True,
                num_rows="fixed", column_config=col_cfg, key=f"pfe_ed_{tipo}_{fkey}",
            )

            for excel_row, val in zip(df_f["_row"].tolist(), edit[COL_EDIT_LABEL].tolist()):
                v = pd.to_numeric(val, errors="coerce")
                store[int(excel_row)] = None if pd.isna(v) else float(v)

        total_folha = sum(v for v in store.values() if v is not None)
        st.caption(f"Σ Nº Ações a desenvolver ({tipo}, folha completa): **{total_folha:.0f}** ações")

    # Gerar o ficheiro: aplicar o store à coluna alvo de cada folha, preservando o resto.
    wb = load_workbook(BytesIO(file_bytes))
    for tipo, nome_folha in folhas_ano.items():
        ws = wb[nome_folha]
        c_alvo = _idx_col_alvo_openpyxl(ws)
        if c_alvo is None:
            continue
        store = st.session_state.pfe_alvo.get(nome_folha, {})
        for excel_row, val in store.items():
            cell = ws.cell(row=int(excel_row), column=c_alvo)
            if val is None or pd.isna(val):
                cell.value = None
            else:
                fv = float(val)
                cell.value = int(fv) if fv.is_integer() else fv

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    st.markdown("---")
    st.download_button(
        "📥 Descarregar projeção anual editada", data=out.getvalue(),
        file_name=f"PFE_{ano}_editado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — Análise da reprojeção do 2º semestre (SÓ LEITURA, sem download)
# ════════════════════════════════════════════════════════════════════════════
def _reforecast_snapshot():
    st.subheader("📁 Carregar snapshot do 1º semestre")

    snap_file = st.file_uploader(
        "Carregue o ficheiro 'reforecast_AAAA-MM-DD.xlsx' exportado na página de Qualidade",
        type=["xlsx", "xls"],
        key="upload_snapshot_reforecast",
    )
    if snap_file is None:
        st.info("⬆️ Carregue o snapshot para analisar a situação a meio do ano.")
        return

    try:
        df = pd.read_excel(snap_file)
    except Exception as e:
        st.error(f"Erro ao ler o ficheiro: {e}")
        return

    obrigatorias = ['Centro', 'Código curso', 'Previstas', 'Finalizadas']
    em_falta = [c for c in obrigatorias if c not in df.columns]
    if em_falta:
        st.error(f"Colunas obrigatórias em falta: {em_falta}")
        st.info(f"Colunas encontradas: {list(df.columns)}")
        return

    df['Previstas'] = pd.to_numeric(df['Previstas'], errors='coerce').fillna(0)
    df['Finalizadas'] = pd.to_numeric(df['Finalizadas'], errors='coerce').fillna(0)

    data_retrato = str(df['Data Retrato'].iloc[0]) if 'Data Retrato' in df.columns and len(df) else "—"
    st.success(f"✅ Snapshot carregado: **{len(df)}** linhas | Data do retrato: **{data_retrato}**")
    st.caption("ℹ️ Esta página é só de leitura. A edição faz-se no separador **Editar Projeção**.")

    df['Falta p/ Plano'] = (df['Previstas'] - df['Finalizadas']).clip(lower=0).astype(int)
    df['% Cumprido (1º sem)'] = 0.0
    m = df['Previstas'] > 0
    df.loc[m, '% Cumprido (1º sem)'] = (df.loc[m, 'Finalizadas'] / df.loc[m, 'Previstas'] * 100).round(1)

    st.markdown("---")
    centros = sorted(df['Centro'].dropna().astype(str).unique())
    sel = st.multiselect(
        "🏢 Filtrar por Centro", centros, default=centros, key="reforecast_centro_filter",
    )
    filt = df[df['Centro'].astype(str).isin(sel)] if sel else df
    parcial = bool(sel) and len(sel) < len(centros)

    if filt.empty:
        st.info("Nenhuma linha para os centros selecionados.")
        return

    prev = filt['Previstas'].sum()
    fin = filt['Finalizadas'].sum()
    st.subheader("📊 Situação a meio do ano" + (f" — {len(sel)} centro(s)" if parcial else " (todos os centros)"))
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("Plano (denominador)", f"{prev:.0f}")
    with c2:
        st.metric("Finalizado (1º sem)", f"{fin:.0f}")
    with c3:
        st.metric("Falta p/ plano", f"{max(prev - fin, 0):.0f}")
    with c4:
        st.metric("% Cumprido (1º sem)", f"{(fin / prev * 100) if prev > 0 else 0:.1f}%")
    if parcial:
        glob_prev = df['Previstas'].sum()
        glob_fin = df['Finalizadas'].sum()
        st.caption(
            f"🌍 Global (todos os centros): {glob_fin:.0f}/{glob_prev:.0f} "
            f"= {(glob_fin / glob_prev * 100) if glob_prev > 0 else 0:.1f}% cumprido"
        )

    cols_final = ['Centro', 'Código curso', 'Previstas', 'Finalizadas',
                  'Falta p/ Plano', '% Cumprido (1º sem)']
    st.dataframe(
        filt[cols_final], use_container_width=True, hide_index=True,
        column_config={
            "% Cumprido (1º sem)": st.column_config.NumberColumn("% Cumprido (1º sem)", format="%.1f%%"),
        },
    )


def mostrar_reforecast():
    st.header("🔄 Reprojeção do 2º Semestre")
    tab1, tab2 = st.tabs(["✏️ Editar Projeção", "📊 Análise 2º Semestre"])
    with tab1:
        _editor_pfe()
    with tab2:
        _reforecast_snapshot()


if __name__ == "__main__":
    mostrar_reforecast()