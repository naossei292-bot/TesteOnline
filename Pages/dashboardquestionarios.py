import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io
import tempfile
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.drawing.image import Image as XLImage
import plotly.io as pio

# ─────────────────────────────────────────────────────────────
# Funções de exportação (mantidas iguais)
# ─────────────────────────────────────────────────────────────

def exportar_dashboard_para_excel(df_filtrado, df_centro, df_resp, df_plot_folha, titulo_arquivo="dashboard_completo.xlsx"):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not df_filtrado.empty:
            df_filtrado.to_excel(writer, sheet_name='Dados Filtrados', index=False)
            worksheet = writer.sheets['Dados Filtrados']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        if not df_centro.empty:
            df_centro.to_excel(writer, sheet_name='Satisfação por Centro', index=False)
        if not df_resp.empty:
            df_resp.to_excel(writer, sheet_name='Satisfação por Respondente', index=False)
        if not df_plot_folha.empty:
            df_plot_folha.to_excel(writer, sheet_name='Satisfação por Folha', index=False)
    output.seek(0)
    return output

def exportar_dashboard_completo(df_filtrado, df_centro, df_resp, df_plot_folha, 
                                 fig_centro, fig_resp, fig_folha, fig_evolucao):
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    import tempfile
    import os
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if df_filtrado is not None and not df_filtrado.empty:
            df_filtrado.to_excel(writer, sheet_name='1. Dados Filtrados', index=False)
        if df_centro is not None and not df_centro.empty:
            df_centro.to_excel(writer, sheet_name='2. Centro - Dados', index=False)
        if df_resp is not None and not df_resp.empty:
            df_resp.to_excel(writer, sheet_name='3. Respondente - Dados', index=False)
        if df_plot_folha is not None and not df_plot_folha.empty:
            df_plot_folha.to_excel(writer, sheet_name='4. Folha - Dados', index=False)
    
    wb = load_workbook(output)
    temp_images = []
    
    def add_grafico_aba(fig, sheet_name):
        if fig is None:
            return
        temp_img = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
        temp_img.close()
        try:
            fig.write_image(temp_img.name, width=900, height=600, scale=2)
            temp_images.append(temp_img.name)
            ws = wb.create_sheet(sheet_name)
            titulo = fig.layout.title.text if fig.layout and fig.layout.title else sheet_name
            ws['A1'] = titulo
            ws['A1'].font = Font(bold=True, size=14)
            img = XLImage(temp_img.name)
            img.width = 900
            img.height = 600
            ws.add_image(img, 'A3')
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 20
        except Exception as e:
            print(f"Erro ao adicionar gráfico {sheet_name}: {e}")
    
    add_grafico_aba(fig_centro, "📊 Gráfico - Centro")
    add_grafico_aba(fig_resp, "📊 Gráfico - Respondente")
    add_grafico_aba(fig_folha, "📊 Gráfico - Folha")
    add_grafico_aba(fig_evolucao, "📊 Gráfico - Evolução")
    
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    for temp_file in temp_images:
        try:
            os.unlink(temp_file)
        except:
            pass
    
    return final_output

# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def _preparar_dados(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Valor Médio"] = pd.to_numeric(df["Valor Médio"], errors="coerce")
    for col in ["Datini", "Datfim", "Data"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    if "Datini" in df.columns:
        df["_Mês"] = df["Datini"].dt.to_period("M").astype(str)
    elif "Data" in df.columns:
        df["_Mês"] = df["Data"].dt.to_period("M").astype(str)
    else:
        df["_Mês"] = None
    for col in ["Centro", "Shortname", "Respondente", "Módulo", "Folha"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().replace({"nan": None, "None": None, "": None})
    if "Shortname" in df.columns:
        df["Shortname"] = df["Shortname"].fillna("Sem Ação")
    return df

CORES_PALETA = ["#1F77B4", "#FF7F0E", "#2CA02C", "#D62728", "#9467BD", "#8C564B", "#E377C2", "#7F7F7F", "#BCBD22", "#17BECF"]

def _fig_bar(df_group, x, y, titulo, horizontal=False):
    if horizontal:
        fig = px.bar(df_group, x=y, y=x, orientation="h", text_auto=".2f", color_discrete_sequence=CORES_PALETA)
        fig.update_traces(textposition="outside", marker_color=CORES_PALETA[0])
    else:
        fig = px.bar(df_group, x=x, y=y, text_auto=".2f", color_discrete_sequence=CORES_PALETA)
        fig.update_traces(textposition="outside", marker_color=CORES_PALETA[0])
    fig.update_layout(title_text=titulo, title_font_size=13, title_font_color="#1F4E79")
    fig.update_yaxes(showgrid=True, gridcolor="#E5E5E5")
    fig.update_xaxes(showgrid=False)
    return fig

def _fig_pizza(df_group, names, values, titulo):
    fig = px.pie(df_group, names=names, values=values, hole=0.45, color_discrete_sequence=CORES_PALETA)
    fig.update_traces(textinfo="percent+label", textfont_size=11, marker=dict(line=dict(color='white', width=2)))
    fig.update_layout(title_text=titulo, title_font_size=13, title_font_color="#1F4E79", showlegend=True, legend=dict(orientation="h", yanchor="bottom", y=-0.1, xanchor="center", x=0.5))
    return fig

def _fig_linha_pontos(df_group, x, y, titulo=""):
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=df_group[x], y=df_group[y], mode='lines+markers+text', name='Série1',
        line=dict(width=2.5, color=CORES_PALETA[0]),
        marker=dict(size=12, color=CORES_PALETA[0], symbol='circle', line=dict(color='white', width=2)),
        text=df_group[y].round(1), textposition='top center', textfont=dict(size=10, color="#1F4E79")
    ))
    fig.update_layout(
        title=dict(text=titulo, font_size=13, font_color="#1F4E79"),
        xaxis=dict(title=dict(text="", font_size=11), tickangle=-45 if len(df_group) > 10 else 0, showgrid=False, tickfont=dict(size=10)),
        yaxis=dict(title=dict(text="Valor Médio", font_size=11), showgrid=True, gridcolor="#E5E5E5", range=[0.5, 4.5], dtick=0.5, tickfont=dict(size=10)),
        plot_bgcolor='white', paper_bgcolor='white'
    )
    return fig

def _fig_barras_por_folha(df_group, categoria=None, titulo=""):
    ordem_folhas = ["21a", "21b", "22a", "22b", "23a", "23b", "24a", "24b"]
    if categoria and categoria in df_group.columns:
        fig = px.bar(df_group.sort_values(["Folha", categoria]), x="Folha", y="Valor Médio", color=categoria, barmode="group", category_orders={"Folha": ordem_folhas}, color_discrete_sequence=CORES_PALETA)
        fig.update_traces(texttemplate="%{y:.2f}", textposition="outside")
    else:
        fig = px.bar(df_group.sort_values("Folha"), x="Folha", y="Valor Médio", category_orders={"Folha": ordem_folhas}, color_discrete_sequence=[CORES_PALETA[0]])
        fig.update_traces(texttemplate="%{y:.2f}", textposition="outside", marker_color=CORES_PALETA[0])
    fig.update_layout(title_text=titulo, title_font_size=13, title_font_color="#1F4E79", barmode="group", plot_bgcolor='white', paper_bgcolor='white')
    fig.update_yaxes(showgrid=True, gridcolor="#E5E5E5", range=[0.5, 4.5], dtick=0.5)
    fig.update_xaxes(showgrid=False, type="category")
    return fig

# ─────────────────────────────────────────────────────────────
# Filtros (AGORA FORA DO EXPANDER - MAIS RÁPIDO)
# ─────────────────────────────────────────────────────────────

def _aplicar_filtros(df: pd.DataFrame, sel_respondente, sel_centro, sel_shortname, sel_modulo, sel_data_ini, sel_data_fim) -> pd.DataFrame:
    df_f = df.copy()
    if sel_respondente:
        df_f = df_f[df_f["Respondente"].isin(sel_respondente)]
    if sel_centro:
        df_f = df_f[df_f["Centro"].isin(sel_centro)]
    if sel_shortname:
        df_f = df_f[df_f["Shortname"].isin(sel_shortname)]
    if sel_modulo:
        df_f = df_f[df_f["Módulo"].isin(sel_modulo)]
    if sel_data_ini and sel_data_fim and "Datini" in df_f.columns:
        mask = (df_f["Datini"].isna() | ((df_f["Datini"].dt.date >= sel_data_ini) & (df_f["Datini"].dt.date <= sel_data_fim)))
        df_f = df_f[mask]
    return df_f

def mostrar_filtros_rapidos(df: pd.DataFrame):
    """
    Filtros fora do expansor para carregamento mais rápido
    """
    st.markdown("### 🔍 Filtros Rápidos")
    
    # Inicializar session state
    if 'filtros_respondente' not in st.session_state:
        st.session_state.filtros_respondente = []
    if 'filtros_centro' not in st.session_state:
        st.session_state.filtros_centro = []
    if 'filtros_shortname' not in st.session_state:
        st.session_state.filtros_shortname = []
    if 'filtros_modulo' not in st.session_state:
        st.session_state.filtros_modulo = []
    if 'filtros_data_ini' not in st.session_state:
        st.session_state.filtros_data_ini = None
    if 'filtros_data_fim' not in st.session_state:
        st.session_state.filtros_data_fim = None
    
    # Linha 1: Respondente e Centro
    col1, col2 = st.columns(2)
    
    with col1:
        todos_respondentes = sorted(df["Respondente"].dropna().unique())
        sel_respondente = st.multiselect("Respondente", todos_respondentes, default=st.session_state.filtros_respondente, key="filtro_respondente_rapido", placeholder="Selecione...")
        st.session_state.filtros_respondente = sel_respondente
    
    # Filtrar baseado em respondente
    df_temp = df.copy()
    if sel_respondente:
        df_temp = df_temp[df_temp["Respondente"].isin(sel_respondente)]
    
    with col2:
        todos_centros = sorted(df_temp["Centro"].dropna().unique())
        sel_centro = st.multiselect("Centro", todos_centros, default=[c for c in st.session_state.filtros_centro if c in todos_centros], key="filtro_centro_rapido", placeholder="Selecione...")
        st.session_state.filtros_centro = sel_centro
    
    # Filtrar por centro
    if sel_centro:
        df_temp = df_temp[df_temp["Centro"].isin(sel_centro)]
    
    # Linha 2: Ação/Shortname e Módulo
    col3, col4 = st.columns(2)
    
    with col3:
        todos_shortnames = sorted(df_temp["Shortname"].dropna().unique())
        sel_shortname = st.multiselect("Ação / Shortname", todos_shortnames, default=[s for s in st.session_state.filtros_shortname if s in todos_shortnames], key="filtro_shortname_rapido", placeholder="Selecione...")
        st.session_state.filtros_shortname = sel_shortname
    
    if sel_shortname:
        df_temp = df_temp[df_temp["Shortname"].isin(sel_shortname)]
    
    with col4:
        todos_modulos = sorted(df_temp["Módulo"].dropna().unique())
        sel_modulo = st.multiselect("Módulo", todos_modulos, default=[m for m in st.session_state.filtros_modulo if m in todos_modulos], key="filtro_modulo_rapido", placeholder="Selecione...")
        st.session_state.filtros_modulo = sel_modulo
    
    if sel_modulo:
        df_temp = df_temp[df_temp["Módulo"].isin(sel_modulo)]
    
    # Linha 3: Datas
    col5, col6 = st.columns(2)
    
    with col5:
        datas_validas = df_temp["Datini"].dropna()
        if not datas_validas.empty:
            d_min = datas_validas.min().date()
            d_max = datas_validas.max().date()
            sel_data_ini = st.date_input("Data Início (De)", value=st.session_state.filtros_data_ini if st.session_state.filtros_data_ini else d_min, min_value=d_min, max_value=d_max, key="filtro_data_ini_rapido")
            st.session_state.filtros_data_ini = sel_data_ini
        else:
            sel_data_ini = None
    
    with col6:
        if not datas_validas.empty:
            sel_data_fim = st.date_input("Data Início (Até)", value=st.session_state.filtros_data_fim if st.session_state.filtros_data_fim else d_max, min_value=d_min, max_value=d_max, key="filtro_data_fim_rapido")
            st.session_state.filtros_data_fim = sel_data_fim
        else:
            sel_data_fim = None
    
    # Botão de limpar filtros
    if st.button("🗑️ Limpar todos os filtros", use_container_width=True):
        st.session_state.filtros_respondente = []
        st.session_state.filtros_centro = []
        st.session_state.filtros_shortname = []
        st.session_state.filtros_modulo = []
        st.session_state.filtros_data_ini = None
        st.session_state.filtros_data_fim = None
        st.rerun()
    
    # Mostrar contagem
    st.info(f"📊 **{len(df_temp):,}** registos disponíveis com os filtros atuais.", icon="ℹ️")
    
    return (sel_respondente, sel_centro, sel_shortname, sel_modulo, sel_data_ini, sel_data_fim)


# ─────────────────────────────────────────────────────────────
# Página principal da dashboard
# ─────────────────────────────────────────────────────────────

def mostrar_questionarios_dashboard():
    st.set_page_config(layout="wide", page_title="Dashboard de Satisfação")
    
    st.title("📊 Dashboard de Satisfação")
    st.markdown("Análise de questionários por Centro, Ação, Data e Respondente")
    st.markdown("---")

    # Upload de ficheiro (opcional)
    with st.expander("📂 **Carregar ficheiro de dados OPCIONAL**", expanded=False):
        uploaded_file = st.file_uploader("Carregue o ficheiro exportado do módulo Questionários (Excel ou CSV)", type=["xlsx", "csv"], key="dash_upload")

    # Carregar dados
    if uploaded_file is not None:
        if uploaded_file.name.endswith(".xlsx"):
            df_raw = pd.read_excel(uploaded_file)
        else:
            df_raw = pd.read_csv(uploaded_file)
        st.success(f"✅ Dados carregados do ficheiro: {len(df_raw)} registos")
        # Resetar filtros
        st.session_state.filtros_respondente = []
        st.session_state.filtros_centro = []
        st.session_state.filtros_shortname = []
        st.session_state.filtros_modulo = []
        st.session_state.filtros_data_ini = None
        st.session_state.filtros_data_fim = None
    elif "quest_editaveis" in st.session_state and not st.session_state.quest_editaveis.empty:
        df_raw = st.session_state.quest_editaveis.drop(columns=["Apagar"], errors="ignore").copy()
    else:
        st.warning("⚠️ Sem dados. Carregue um ficheiro exportado da página Questionários ou aceda a essa página e carregue os ficheiros primeiro.", icon="📋")
        return

    # Preparar dados
    df_raw = _preparar_dados(df_raw)
    if df_raw.empty:
        st.info("ℹ️ Nenhum dado válido encontrado.")
        return

    # Filtros rápidos (FORA DO EXPANDER)
    sel_respondente, sel_centro, sel_shortname, sel_modulo, sel_data_ini, sel_data_fim = mostrar_filtros_rapidos(df_raw)

    # Aplicar filtros
    df = _aplicar_filtros(df_raw, sel_respondente, sel_centro, sel_shortname, sel_modulo, sel_data_ini, sel_data_fim)

    if df.empty:
        st.info("ℹ️ Nenhum registo corresponde aos filtros selecionados.")
        return

    st.markdown("---")

    # KPIs
    st.markdown("### 📈 Indicadores Gerais")
    n_total = len(df_raw)
    n_filtrado = len(df)
    vm_geral = df["Valor Médio"].mean()
    vm_max = df.groupby("Shortname")["Valor Médio"].mean().max() if not df.empty else 0
    vm_min = df.groupby("Shortname")["Valor Médio"].mean().min() if not df.empty else 0

    k1, k2, k3, k4, k5 = st.columns(5)
    with k1:
        st.metric("Registos filtrados", f"{n_filtrado:,}".replace(",", "."), f"de {n_total:,}".replace(",", "."))
    with k2:
        st.metric("Cursos (Shortnames)", df["Shortname"].nunique())
    with k3:
        st.metric("Centros", df["Centro"].nunique() if df["Centro"].notna().any() else "—")
    with k4:
        st.metric("Média geral", f"{vm_geral:.2f}" if pd.notna(vm_geral) else "—")
    with k5:
        st.metric("Média max/min", f"{vm_max:.2f}" if pd.notna(vm_max) else "—", f"min: {vm_min:.2f}" if pd.notna(vm_min) else "—")

    st.markdown("")

    # Gráficos
    st.markdown("### 📊 Satisfação por Centro e Respondente")
    c1, c2 = st.columns([3, 2])

    with c1:
        if df["Centro"].notna().any():
            df_centro = df.groupby("Centro")["Valor Médio"].mean().round(2).reset_index().sort_values("Valor Médio", ascending=True).rename(columns={"Valor Médio": "Média"})
            st.plotly_chart(_fig_bar(df_centro, "Centro", "Média", "Valor Médio por Centro", horizontal=True), use_container_width=True)
        else:
            st.info("Carregue o ficheiro Excel de Ações para ver dados por Centro.")

    with c2:
        if df["Respondente"].notna().any():
            df_resp = df.groupby("Respondente")["Valor Médio"].mean().round(2).reset_index().rename(columns={"Valor Médio": "Média"})
            st.plotly_chart(_fig_pizza(df_resp, "Respondente", "Média", "Satisfação Média por Respondente"), use_container_width=True)

    # Gráfico de barras por folha
    st.markdown("### 📈 Evolução da Satisfação")

    if "Folha" in df.columns and "Valor Médio" in df.columns:
        ordem_folhas = ["21a", "21b", "22a", "22b", "23a", "23b", "24a", "24b"]
        df_folha = df[df["Folha"].isin(ordem_folhas)].copy()
        if not df_folha.empty:
            df_folha["Folha"] = pd.Categorical(df_folha["Folha"], categories=ordem_folhas, ordered=True)
            if "Pergunta" in df_folha.columns and df_folha["Pergunta"].notna().any():
                letras_validas = ["A", "B", "C", "D", "E", "F"]
                df_folha["Pergunta Letra"] = df_folha["Pergunta"].astype(str).str.strip().str.upper().str.extract(r'^([A-Z])')[0]
                df_folha = df_folha[df_folha["Pergunta Letra"].isin(letras_validas)]
                df_plot = df_folha.groupby(["Folha", "Pergunta Letra"])["Valor Médio"].mean().round(2).reset_index().rename(columns={"Pergunta Letra": "Pergunta"})
                st.plotly_chart(_fig_barras_por_folha(df_plot, categoria="Pergunta", titulo="Satisfação Média por Folha e Pergunta"), use_container_width=True)
            else:
                df_plot = df_folha.groupby("Folha")["Valor Médio"].mean().round(2).reset_index()
                st.plotly_chart(_fig_barras_por_folha(df_plot, categoria=None, titulo="Satisfação Média por Folha"), use_container_width=True)
        else:
            st.info("ℹ️ Não existem folhas 21a-24b nos dados atuais para este gráfico.")
    elif "_Mês" in df.columns and "Valor Médio" in df.columns:
        df_meses = df.groupby("_Mês")["Valor Médio"].mean().round(2).reset_index().sort_values("_Mês")
        st.plotly_chart(_fig_linha_pontos(df_meses, "_Mês", "Valor Médio", "Evolução Mensal da Satisfação"), use_container_width=True)
    elif "Shortname" in df.columns and "Valor Médio" in df.columns:
        df_acoes = df.groupby("Shortname")["Valor Médio"].mean().round(2).reset_index().sort_values("Shortname")
        st.plotly_chart(_fig_linha_pontos(df_acoes, "Shortname", "Valor Médio", "Satisfação Média por Ação"), use_container_width=True)
    else:
        st.info("ℹ️ Adicione uma coluna 'Pergunta', 'Shortname' ou tenha dados mensais para ver o gráfico de evolução.")

    # Tabela resumo
    st.markdown("### 📋 Tabela Resumo")
    COLS_RESUMO = [c for c in ["Centro", "Shortname", "Datini", "Respondente", "Módulo", "Folha", "Pergunta", "Valor Médio"] if c in df.columns]
    df_resumo = df[COLS_RESUMO].copy()
    if "Datini" in df_resumo.columns:
        df_resumo["Datini"] = df_resumo["Datini"].dt.strftime("%d/%m/%Y")

    st.dataframe(df_resumo.reset_index(drop=True), use_container_width=True, height=280, hide_index=True)
    st.caption(f"A mostrar {len(df_resumo):,} registos filtrados.".replace(",", "."))

    # Exportação
    st.markdown("---")
    st.markdown("### 📥 Exportar Dashboard Completo")
    
    # Preparar dados para exportação
    df_centro_export = None
    if df["Centro"].notna().any():
        df_centro_export = df.groupby("Centro")["Valor Médio"].mean().round(2).reset_index().sort_values("Valor Médio", ascending=True).rename(columns={"Valor Médio": "Média"})
    
    df_resp_export = None
    if df["Respondente"].notna().any():
        df_resp_export = df.groupby("Respondente")["Valor Médio"].mean().round(2).reset_index().rename(columns={"Valor Médio": "Média"})
    
    df_folha_export = None
    if "Folha" in df.columns:
        ordem_folhas = ["21a", "21b", "22a", "22b", "23a", "23b", "24a", "24b"]
        df_folha_temp = df[df["Folha"].isin(ordem_folhas)].copy()
        if not df_folha_temp.empty:
            df_folha_export = df_folha_temp.groupby("Folha")["Valor Médio"].mean().round(2).reset_index()
    
    df_resumo_export = df[COLS_RESUMO].copy()
    if "Datini" in df_resumo_export.columns:
        df_resumo_export["Datini"] = df_resumo_export["Datini"].dt.strftime("%d/%m/%Y")
    
    # Gráficos para exportação
    fig_centro_export = _fig_bar(df_centro_export, "Centro", "Média", "Valor Médio por Centro", horizontal=True) if df_centro_export is not None else None
    fig_resp_export = _fig_pizza(df_resp_export, "Respondente", "Média", "Satisfação Média por Respondente") if df_resp_export is not None else None
    fig_folha_export = _fig_barras_por_folha(df_folha_export, categoria=None, titulo="Satisfação Média por Folha") if df_folha_export is not None else None
    
    col_export1, col_export2 = st.columns([1, 2])
    
    with col_export1:
        if st.button("📊 Exportar Dashboard Completo (Excel)", use_container_width=True, type="primary"):
            with st.spinner("A gerar ficheiro Excel..."):
                excel_data = exportar_dashboard_completo(df_resumo_export, df_centro_export, df_resp_export, df_folha_export, fig_centro_export, fig_resp_export, fig_folha_export, None)
                st.download_button(label="✅ Clique para descarregar", data=excel_data, file_name=f"dashboard_satisfacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="download_dashboard_completo")
    
    with col_export2:
        num_centros = len(df_centro_export) if df_centro_export is not None and not df_centro_export.empty else 0
        st.caption(f"📌 A exportar dados de {len(df_resumo_export):,} registos e {num_centros} centros")

if __name__ == "__main__":
    mostrar_questionarios_dashboard()