import hashlib
import io
import re

import pandas as pd
import streamlit as st

# ─────────────────────────────────────────────────────────────
# Parsing helpers
# ─────────────────────────────────────────────────────────────

RE_ITEM = re.compile(r'^(?:(M\d+)_)?([^.]+)\.(.+)$')

FOLHA_META = {
    "21a": {"Respondente": "Formando",               "Modalidade": "Presencial",  "Tipo": "Módulo"},
    "21b": {"Respondente": "Formando",               "Modalidade": "À Distância", "Tipo": "Módulo"},
    "22a": {"Respondente": "Formando",               "Modalidade": "Presencial",  "Tipo": "Ação"},
    "22b": {"Respondente": "Formando",               "Modalidade": "À Distância", "Tipo": "Ação"},
    "23a": {"Respondente": "Formador",               "Modalidade": "Presencial",  "Tipo": "Ação"},
    "23b": {"Respondente": "Tutor",                  "Modalidade": "À Distância", "Tipo": "Ação"},
    "24a": {"Respondente": "Coordenação Pedagógica", "Modalidade": "Presencial",  "Tipo": "Formador"},
    "24b": {"Respondente": "Coordenação Pedagógica", "Modalidade": "À Distância", "Tipo": "Tutor"},
}
# ── Dicionário com as perguntas completas para cada folha ──────────────────
PERGUNTAS_POR_FOLHA = {
    "21a": {
        "A01": "A01-Coerência de estruturação dos conteúdos.",
        "A02": "A02-Relevância dos temas abordados.",
        "A03": "A03-Adequação da duração do Módulo.",
        "A04": "A04-Nível de abrangência da documentação disponibilizada.",
        "A05": "A05-Condições de limpeza da sala de formação utilizada.",
        "B01": "B01-Apoio logístico e administrativo prestado pela Equipa de Coordenação Pedagógica de Centro (entrega de material, exames, contratos, resolução de problemas informáticos).",
        "C01": "C01-Correspondência entre os conteúdos do módulo e as suas expectativas iniciais",
        "C02": "C02-Aplicação dos conteúdos no campo profissional e/ou pessoal.",
        "D01": "D01-O Formador domina os temas abordados.",
        "D02": "D02-O formador domína as técnicas pedagógicas.",
        "D03": "D03-O formador motiva os formandos.",
        "D04": "D04-O formador demonstra capacidade de exposição dos conteúdos.",
        "D05": "D05-O formador utiliza adequadamente os procedimentos didáticos (aulas expositivas, aulas práticas, textos de apoio, trabalhos de grupo, atividades externas).",
        "D06": "D06-O formador demonstra disponibilidade para esclarecer as questões dos formandos.",
        "E01": "E01-Avaliação global do módulo frequentado.",
    },
    "21b": {
        "A01": "A01-Qualidade dos recursos didáticos.",
        "A02": "A02-Coerência da estruturação dos conteúdos.",
        "A03": "A03-Relevância dos temas abordados.",
        "A04": "A04-Nível de abrangência da documentação distribuída.",
        "A05": "A05-Adequação da duração do módulo.",
        "A06": "A06-Adequação da forma de organização.",
        "B01": "B01-Apoio prestado pela Coordenação Pedagógica.",
        "B02": "B02-Apoio administrativo prestado.",
        "B03": "B03-Apoio técnico de suporte.",
        "C01": "C01-Correspondência entre os conteúdos do módulo e as suas expectativas iniciais.",
        "C02": "C02-Aplicação dos conteúdos no campo profissional e/ou pessoal.",
        "D01": "D01-O Formador/Tutor domina os temas abordados.",
        "D02": "D02-O Formador/Tutor demonstra disponibilidade para esclarecer as questões dos formandos.",
        "E01": "E01-Acessibilidade e navegabilidade.",
        "E02": "E02-Qualidade gráfica dos documentos.",
        "E03": "E03-Qualidade do sistema em formação à distância.",
        "F01": "F01-Avaliação global do módulo frequentado.",
    },
    "22a": {
        "A01": "A01-Adequação da sala de formação utilizada.",
        "A02": "A02-Qualidade dos recursos didáticos adotados.",
        "A03": "A03-Adequação do desenvolvimento da ação de formação e organização dos módulos ministrados.",
        "A04": "A04-Importância e utilidade dos conteúdos ministrados.",
        "A05": "A05-Cumprimento dos objetivos da ação de formação.",
        "A06": "A06-Número de horas destinados à ação de formação.",
        "B01": "B01-Apresentação preliminar do plano de ensino pela Equipa de Coordenação Pedagógica de Centro (programa, cronograma, avaliação).",
        "B02": "B02-Encerramento da ação de formação pela Equipa de Coordenação Pedagógica de Centro (indicação de progressão, apresentação de estágio, se aplicável, orientações gerais).",
        "B03": "B03-Apoio prestado pela Coordenação Pedagógica de Centro (gestão de conflitos, disponibilidade no atendimento, clareza na comunicação).",
        "C01": "C01-Correspondência entre os conteúdos do módulo e as suas expectativas iniciais",
        "D01": "D01-Avaliação global do desempenho do(s) formador(es).",
        "E01": "E01-Avaliação global da ação de formação frequentada.",
    },
    "22b": {
        "A01": "A01-Qualidade dos recursos didáticos.",
        "A02": "A02-Coerência da estruturação dos conteúdos.",
        "A03": "A03-Relevância dos temas abordados.",
        "A04": "A04-Nível de abrangência da documentação distribuída.",
        "A05": "A05-Adequação da duração da Ação de Formação.",
        "A06": "A06-Adequação da forma de organização.",
        "B01": "B01-Apoio prestado pela Coordenação Pedagógica.",
        "B02": "B02-Apoio administrativo prestado.",
        "B03": "B03-Apoio técnico de suporte.",
        "C01": "C01-Correspondência entre os conteúdos do módulo e as suas expectativas iniciais.",
        "C02": "C02-Aplicação dos conteúdos no campo profissional e/ou pessoal.",
        "D01": "D01-O Formador/Tutor domina os temas abordados.",
        "D02": "D02-O Formador/Tutor demonstra disponibilidade para esclarecer as questões dos formandos.",
        "E01": "E01-Acessibilidade e navegabilidade.",
        "E02": "E02-Qualidade gráfica dos documentos.",
        "E03": "E03-Qualidade do sistema em formação à distância.",
        "F01": "F01-Avaliação global da ação de formação frequentada.",
    },
    "23a": {
        "A01": "A01-Cumprimentos dos Objetivos da ação.",
        "A02": "A02-Coerência da estruturação dos conteúdos ao longo da ação.",
        "A03": "A03-Utilidade dos conteúdos.",
        "A04": "A04-Adequação da duração do curso.",
        "A05": "A05-Qualidade dos recursos didáticos.",
        "A06": "A06-Adequação da sala de formação utilizada.",
        "A07": "A07-Atividades Externas/Recursos Programados.",
        "A08": "A08-Conhecimento prévio do público-alvo.",
        "A09": "A09-Trabalho com a equipa pedagógica (formadores e coordenação).",
        "B01": "B01-Apoio prestado pela Coordenação (apresentação inicial, gestão de conflitos, disponibilidade, clareza na comunicação).",
        "B02": "B02-Apoio logístico e administrativo prestado pela equipa de coordenação do centro de formação (entrega de material, exames, contratos, resolução de problemas informáticos).",
        "C01": "C01-Motivação do grupo.",
        "C02": "C02-Grau de participação do grupo.",
        "C03": "C03-Grau de adequação de postura e perfil profissional demonstrado em sala.",
        "C04": "C04-Disponibilidade do grupo para realizar atividades pedagógicas e produtividade.",
        "C05": "C05-Relacionamento criado entre os formandos.",
        "C06": "C06-Relacionamento criado entre o formador e os formandos.",
        "C07": "C07-Alteração do perfil de entrada e de saída dos formandos.",
        "C08": "C08-Resultados alcançados pelos formandos.",
        "C09": "C09-Assiduidade e Pontualidade do grupo.",
        "D01": "D01-Apreciação global do desempenho do formador.",
    },
    "23b": {
        "A01": "A01-Qualidade dos recursos didáticos.",
        "A02": "A02-Utilidade dos conteúdos.",
        "A03": "A03-Adequação do programa face aos objetivos previstos.",
        "A04": "A04-Coerência da estruturação dos conteúdos ao longo da ação.",
        "A05": "A05-Nível de abrangência da documentação distribuída.",
        "A06": "A06-Adequação da duração do curso.",
        "B01": "B01-Apoio prestado pela Coordenação Pedagógica.",
        "B02": "B02-Apoio logístico e administrativo prestado.",
        "B03": "B03-Apoio prestado pela equipa de suporte técnico.",
        "C01": "C01-Motivação do formando.",
        "C02": "C02-Grau de participação do formando nas atividades propostas.",
        "C03": "C03-Disponibilidade do formando no cumprimento do Itinerário de Atividades.",
        "D01": "D01-Acessibilidade e navegabilidade.",
        "D02": "D02-Qualidade do sistema em e-Learning.",
        "E01": "E01-Avaliação Global do Curso.",
    },
    "24a": {
        "A01": "A01-Preenche corretamente e no tempo adequado os documentos do DTP da sua responsabilidade.",
        "A02": "A02-Planifica corretamente as sessões e as atividades pedagógicas de acordo com o programa de formação.",
        "A03": "A03-Assiduidade.",
        "A04": "A04-Pontualidade.",
        "A05": "A05-Reporta as questões relativas à Coordenação.",
        "A06": "A06-Participa ativamente nas reuniões da Equipa Pedagógica.",
        "B01": "B01-Utiliza métodos e estratégias pedagógicas adequadas ao perfil dos formandos.",
        "B02": "B02-Demonstra preocupação em fazer corresponder os assuntos da formação aos interesses e expectativas dos formandos.",
        "B03": "B03-Fomenta o espírito de iniciativa e de autonomia dos formandos.",
        "B04": "B04-Domina o assunto da formação.",
        "C01": "C01-Estabelece um bom relacionamento com a Equipa Pedagógica.",
        "C02": "C02-Desenvolve um bom relacionamento com o grupo de formandos.",
        "C03": "C03-Apresenta uma atitude positiva nos relacionamentos interpessoais e em situações imprevistas.",
        "C04": "C04-Demonstra disponibilidade nos relacionamentos interpessoais.",
        "D01": "D01-Apreciação global do desempenho do formador.",
    },
    "24b": {
        "A01": "A01-Preenche corretamente e no tempo adequado os documentos do DTP da sua responsabilidade.",
        "A02": "A02-Planifica corretamente as sessões e as atividades pedagógicas de acordo com o programa de formação.",
        "A03": "A03-Produz com qualidade materiais didáticos e atividades para serem utilizados em formação à distância.",
        "A04": "A04-Reporta as questões relativas à Coordenação.",
        "B01": "B01-Utiliza métodos e estratégias pedagógicas adequadas ao perfil dos formandos.",
        "B02": "B02-Demonstra preocupação em fazer corresponder os assuntos da formação aos interesses e expectativas dos formandos.",
        "B03": "B03-Fomenta o espírito de iniciativa e de autonomia dos formandos.",
        "B04": "B04-Domina o assunto da formação.",
        "C01": "C01-Estabelece um bom relacionamento com a Equipa Pedagógica.",
        "C02": "C02-Pratica uma tutoria ativa através da plataforma Moodle.",
        "D01": "D01-Apreciação global do desempenho do/a tutor/a.",
    },
}

def pergunta_completa(folha: str, codigo: str) -> str:
    """Devolve a pergunta completa para uma dada folha e código."""
    folha = str(folha).strip()
    codigo = str(codigo).strip()
    if folha in PERGUNTAS_POR_FOLHA:
        return PERGUNTAS_POR_FOLHA[folha].get(codigo, codigo)
    return codigo  # fallback: mantém o código se a folha for desconhecida

# Lookup plano para evitar .get() com dict aninhado em cada linha
_FOLHA_RESPONDENTE = {k: v["Respondente"] for k, v in FOLHA_META.items()}
_FOLHA_MODALIDADE  = {k: v["Modalidade"]  for k, v in FOLHA_META.items()}
_FOLHA_TIPO        = {k: v["Tipo"]        for k, v in FOLHA_META.items()}


# ── MELHORIA 1: parsear_item vectorizado (substitui apply linha-a-linha) ──────
def parsear_itens_vectorizado(series: pd.Series) -> pd.DataFrame:
    s = series.astype(str).str.strip().str.replace("__", "_", regex=False)
    extracted = s.str.extract(r'^(?:(M\d+)_)?([^.]+)\.(.+)$')
    # --- NOVO: normalização de itens mal formatados ---
    def normalizar_item(item: str) -> str:
        item = str(item).strip()
        # Se já está no formato esperado (ex: M1_21a.A01), mantém
        if re.match(r'^(?:M\d+_)?[0-9]{2}[a-z]\.[A-Z0-9]+$', item):
            return item
        # Se for "CP_210" ou "CL_210", mapeia para folhas conhecidas (ajuste conforme sua realidade)
        if item.startswith("CP_210"):
            return "24a.CP"       # exemplo: assume folha 24a, pergunta genérica
        if item.startswith("CL_210"):
            return "24b.CL"       # exemplo: folha 24b
        # Se for só números (ex: "244", "235"), atribui a uma folha padrão (ex: 21a)
        if item.isdigit():
            return f"21a.{item}"
        # Para outros casos, tenta extrair qualquer coisa antes do último ponto (fallback)
        if '.' in item:
            return item
        # Último recurso: coloca na folha 21a com o valor original como código
        return f"21a.{item}"
    
    s_normalizado = s.apply(normalizar_item)
    extracted = s_normalizado.str.extract(r'^(?:(M\d+)_)?([^.]+)\.(.+)$')
    extracted.columns = ["Módulo", "Folha", "Pergunta_Codigo"]
    folha = extracted["Folha"]
    extracted["Respondente"] = folha.map(_FOLHA_RESPONDENTE)
    extracted["Modalidade"]  = folha.map(_FOLHA_MODALIDADE)
    extracted["Tipo"]        = folha.map(_FOLHA_TIPO)

    # --- NOVO: criar coluna com a pergunta completa ---
    extracted["Pergunta"] = extracted.apply(
        lambda row: pergunta_completa(row["Folha"], row["Pergunta_Codigo"]),
        axis=1
    )
    return extracted


# ── MELHORIA 2: formatar datas vectorizado (substitui apply linha-a-linha) ────
def formatar_coluna_data(col: pd.Series) -> pd.Series:
    """
    Formata uma coluna de datas para dd/mm/aaaa de forma vectorizada.
    Valores não convertíveis (ex: "FINALIZADA") são mantidos como estão.
    """
    convertida = pd.to_datetime(col, errors="coerce")
    formatada  = convertida.dt.strftime("%d/%m/%Y")
    # Onde não converteu, mantém o valor original como string
    original   = col.where(col.isna(), col.astype(str))
    return formatada.where(convertida.notna(), other=original)


# ── MELHORIA 3: hash rápido para comparação de DataFrames ─────────────────────
def hash_df(df: pd.DataFrame) -> str:
    """MD5 sobre os hashes das linhas — muito mais rápido do que .astype(str).equals()."""
    try:
        return hashlib.md5(
            pd.util.hash_pandas_object(df, index=True).values
        ).hexdigest()
    except Exception:
        # Fallback seguro se o df tiver tipos não hasháveis
        return hashlib.md5(
            df.fillna("__NA__").astype(str).to_csv(index=False).encode()
        ).hexdigest()


# ─────────────────────────────────────────────────────────────
# Leitura de ficheiros (com @st.cache_data)
# ─────────────────────────────────────────────────────────────

# ── MELHORIA 5: cache_data — evita reprocessamento em reruns acidentais ───────
@st.cache_data(show_spinner="A processar Excel de Ações...")
def processar_acoes_xlsx(conteudo: bytes) -> pd.DataFrame:
    """
    Lê ações_por_centro.xlsx → DataFrame com Ação, Datini, Datfim, Centro, U_status.
    Recebe bytes (em vez de file object) para ser compatível com @st.cache_data.
    """
    import unicodedata

    df = pd.read_excel(io.BytesIO(conteudo))
    df.columns = df.columns.str.strip()

    # ── MELHORIA 4: strip vectorizado por coluna object (sem loop Python) ─────
    obj_cols = df.select_dtypes(include="object").columns
    df[obj_cols] = df[obj_cols].apply(lambda c: c.str.strip())

    def norm(s):
        return unicodedata.normalize("NFC", str(s)).lower().strip()

    rename = {}
    for col in df.columns:
        cn = norm(col)
        if cn in ("ação", "acao"):
            rename[col] = "Ação"
        elif cn == "datini":
            rename[col] = "Datini"
        elif cn == "datfim":
            rename[col] = "Datfim"
        elif cn == "deslocal":
            rename[col] = "Centro"
        elif cn in ("u_status", "ustatus", "status"):
            rename[col] = "U_status"

    df = df.rename(columns=rename)

    colunas_necessarias = ["Ação", "Datini", "Datfim", "Centro", "U_status"]
    for col in colunas_necessarias:
        if col not in df.columns:
            df[col] = None

    df["_acao_key"] = df["Ação"].astype(str).str.strip().str.lower()

    return df[colunas_necessarias + ["_acao_key"]].copy()


@st.cache_data(show_spinner="A processar CSV de Questionários...")
def processar_csv(conteudo: bytes) -> pd.DataFrame:
    """
    Lê mdl_course.csv → DataFrame com colunas expandidas e chave normalizada.
    Recebe bytes para ser compatível com @st.cache_data.
    """
    try:
        df = pd.read_csv(io.BytesIO(conteudo), sep=";", encoding="utf-8")
    except UnicodeDecodeError:
        df = pd.read_csv(io.BytesIO(conteudo), sep=";", encoding="latin-1")

    df.columns = df.columns.str.strip()

    # ── MELHORIA 1 aplicada: parsear todos os itens de uma vez (vectorizado) ──
    parsed = parsear_itens_vectorizado(df["item"])
    df = pd.concat([df, parsed], axis=1)

    df["valor_medio"] = pd.to_numeric(df["valor_medio"], errors="coerce")

    df = df.rename(columns={
        "shortname":   "Shortname",
        "data_ini":    "Data",
        "item":        "Item",
        "valor_medio": "Valor Médio",
    })

    df["_shortname_key"] = df["Shortname"].astype(str).str.strip().str.lower()

    return df


def juntar_dados(df_csv: pd.DataFrame, df_acoes: pd.DataFrame) -> pd.DataFrame:
    """
    Faz o join usando as chaves normalizadas (_shortname_key e _acao_key).
    Traz Centro, Datini, Datfim e U_status do Excel.
    """
    if df_acoes.empty:
        df_csv["Centro"]   = None
        df_csv["Datini"]   = None
        df_csv["Datfim"]   = None
        df_csv["U_status"] = None
        return df_csv

    cols_join = ["_acao_key", "Datini", "Datfim", "Centro", "U_status"]
    df_join   = df_acoes[cols_join].rename(columns={"_acao_key": "_shortname_key"})

    df_resultado = df_csv.merge(df_join, on="_shortname_key", how="left")
    df_resultado.drop(columns=["_shortname_key"], inplace=True, errors="ignore")

    return df_resultado


# ─────────────────────────────────────────────────────────────
# Funções auxiliares de estado
# ─────────────────────────────────────────────────────────────

COLUNAS_DADOS = [
    "Shortname", "Centro", "Datini", "Datfim", "U_status",
    "Módulo", "Folha", "Pergunta", "Item",
    "Respondente", "Modalidade", "Tipo",
    "Valor Médio", "Data",
]


def garantir_todas_colunas(df: pd.DataFrame) -> pd.DataFrame:
    for col in COLUNAS_DADOS:
        if col not in df.columns:
            df[col] = None
    if "Apagar" not in df.columns:
        df.insert(0, "Apagar", False)
    cols = ["Apagar"] + [c for c in COLUNAS_DADOS if c in df.columns]
    return df[cols]


def _gerar_excel_com_filtros(df: pd.DataFrame) -> io.BytesIO:
    """Gera um Excel formatado com AutoFilter nativo em todas as colunas."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Questionários"

    COR_CABECALHO = "1F4E79"
    COR_FILTRO    = "2E75B6"
    COLS_DESTAQUE = {"Centro", "Shortname", "Datini", "Datfim", "Respondente"}

    header_base            = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_fill_principal  = PatternFill("solid", fgColor=COR_FILTRO)
    header_fill_normal     = PatternFill("solid", fgColor=COR_CABECALHO)
    align_center           = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left             = Alignment(horizontal="left", vertical="center")
    borda_fina = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    cols = list(df.columns)

    ws.row_dimensions[1].height = 30
    for ci, col_name in enumerate(cols, start=1):
        cell           = ws.cell(row=1, column=ci, value=col_name)
        cell.font      = header_base
        cell.fill      = header_fill_principal if col_name in COLS_DESTAQUE else header_fill_normal
        cell.alignment = align_center
        cell.border    = borda_fina

    fill_par   = PatternFill("solid", fgColor="DCE6F1")
    fill_impar = PatternFill("solid", fgColor="FFFFFF")
    fonte_dado = Font(name="Arial", size=9)

    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        fill = fill_par if ri % 2 == 0 else fill_impar
        for ci, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=ri, column=ci, value=value)
            cell.font      = fonte_dado
            cell.fill      = fill
            cell.alignment = align_left
            cell.border    = borda_fina

    LARGURAS = {
        "Centro": 14, "Shortname": 18, "Datini": 13, "Datfim": 13,
        "Respondente": 22, "Modalidade": 14, "Tipo": 12,
        "Módulo": 10, "Folha": 10, "Pergunta": 12, "Item": 18,
        "Valor Médio": 13, "Data": 13,
    }
    for ci, col_name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = LARGURAS.get(col_name, 16)

    ws.freeze_panes = "A2"

    ultima_col            = get_column_letter(len(cols))
    ws.auto_filter.ref    = f"A1:{ultima_col}{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────
# Página Streamlit
# ─────────────────────────────────────────────────────────────

def mostrar_questionarios():
    st.header("📋 Base de Dados de Satisfação")

    COLUNAS_COM_APAGAR = ["Apagar"] + COLUNAS_DADOS

    # ── Inicializar estado ─────────────────────────────────────
    if "quest_uploader_key"  not in st.session_state:
        st.session_state.quest_uploader_key  = 0
    if "quest_editor_key"    not in st.session_state:
        st.session_state.quest_editor_key    = 0
    if "quest_editaveis"     not in st.session_state:
        st.session_state.quest_editaveis     = pd.DataFrame(columns=COLUNAS_COM_APAGAR)
    if "quest_acoes_df"      not in st.session_state:
        st.session_state.quest_acoes_df      = pd.DataFrame(
            columns=["Ação", "Datini", "Datfim", "Centro", "U_status"]
        )
    if "quest_excel_cache"   not in st.session_state:
        st.session_state.quest_excel_cache   = None
    # ── MELHORIA 3: hash do último export para evitar regeneração ─────────────
    if "quest_excel_hash"    not in st.session_state:
        st.session_state.quest_excel_hash    = None
    # Hash do estado atual dos dados (para detetar mudanças reais)
    if "quest_dados_hash"    not in st.session_state:
        st.session_state.quest_dados_hash    = None

    st.session_state.quest_editaveis = garantir_todas_colunas(
        st.session_state.quest_editaveis.copy()
    )

    # ── Carregar ficheiros ─────────────────────────────────────
    st.subheader("📤 Carregar Ficheiros")

    st.info(
        "**Ficheiro de Ações (Excel):** `ações_por_centro.xlsx` — fornece Ação, Datini, Datfim, Centro.  \n"
        "**Ficheiro de Questionários (CSV):** `mdl_course.csv` — fornece Shortname, Item, Valor Médio.  \n"
        "A correlação é feita automaticamente: **Shortname (CSV) = Ação (Excel)**."
    )

    c1, c2, c3 = st.columns(3)
    with c1:
        modo_carga = st.radio(
            "Modo:", ["Substituir dados existentes", "Adicionar ao final"],
            horizontal=True, key="modo_carga_quest",
        )
    with c2:
        ficheiros = st.file_uploader(
            "Carregar ficheiros (Excel de Ações e/ou CSV de Questionários)",
            type=None,
            accept_multiple_files=True,
            key=f"carga_quest_upload_{st.session_state.quest_uploader_key}",
        )
    with c3:
        if st.button("🗑️ Limpar todos os dados", use_container_width=True, key="btn_limpar_quest"):
            st.session_state.quest_editaveis   = pd.DataFrame(columns=COLUNAS_COM_APAGAR)
            st.session_state.quest_acoes_df    = pd.DataFrame(
                columns=["Ação", "Datini", "Datfim", "Centro", "U_status"]
            )
            st.session_state.quest_excel_cache = None
            st.session_state.quest_excel_hash  = None
            st.session_state.quest_dados_hash  = None
            st.session_state.quest_editor_key += 1
            # Limpar cache das funções para forçar reprocessamento futuro
            processar_csv.clear()
            processar_acoes_xlsx.clear()
            st.rerun()

    if ficheiros:
        dfs_csv   = []
        dfs_acoes = []

        for f in ficheiros:
            nome    = f.name.lower()
            # ── MELHORIA 5: ler bytes uma vez e passar ao cache ───────────────
            conteudo = f.read()
            try:
                if nome.endswith(".xlsx") or nome.endswith(".xls"):
                    df_acao = processar_acoes_xlsx(conteudo)
                    if df_acao.empty:
                        st.warning(f"⚠️ Sem dados extraídos de: {f.name}")
                    else:
                        dfs_acoes.append(df_acao)
                        st.success(f"✅ {f.name} → {len(df_acao)} ações carregadas")

                elif nome.endswith(".csv"):
                    df_csv = processar_csv(conteudo)
                    if df_csv.empty:
                        st.warning(f"⚠️ Sem dados extraídos de: {f.name}")
                    else:
                        dfs_csv.append(df_csv)
                        st.success(
                            f"✅ {f.name} → {len(df_csv)} registos, "
                            f"{df_csv['Shortname'].nunique()} cursos"
                        )
                else:
                    st.warning(f"Formato não suportado: {f.name}")

            except Exception as e:
                st.error(f"Erro em {f.name}: {e}")

        if dfs_acoes:
            df_acoes_novo = pd.concat(dfs_acoes, ignore_index=True).drop_duplicates(subset=["Ação"])
            if modo_carga == "Substituir dados existentes":
                st.session_state.quest_acoes_df = df_acoes_novo
            else:
                st.session_state.quest_acoes_df = pd.concat(
                    [st.session_state.quest_acoes_df, df_acoes_novo], ignore_index=True
                ).drop_duplicates(subset=["Ação"])

            # Re-join se já há CSV carregado
            df_existente = st.session_state.quest_editaveis.drop(
                columns=["Apagar"], errors="ignore"
            )
            if not df_existente.empty and df_existente["Shortname"].notna().any():
                df_csv_atual = df_existente.drop(
                    columns=["Centro", "Datini", "Datfim", "U_status"], errors="ignore"
                )
                df_rejoin = juntar_dados(df_csv_atual, st.session_state.quest_acoes_df)
                df_rejoin = garantir_todas_colunas(df_rejoin)
                df_rejoin["Apagar"] = False
                st.session_state.quest_editaveis = df_rejoin

        if dfs_csv:
            df_csv_total = pd.concat(dfs_csv, ignore_index=True)
            df_joined    = juntar_dados(df_csv_total, st.session_state.quest_acoes_df)
            df_joined    = garantir_todas_colunas(df_joined)
            df_joined["Apagar"] = False

            if modo_carga == "Substituir dados existentes":
                st.session_state.quest_editaveis = df_joined
            else:
                st.session_state.quest_editaveis = pd.concat(
                    [st.session_state.quest_editaveis, df_joined], ignore_index=True
                )

            n_match = df_joined["Centro"].notna().sum()
            n_total = len(df_joined)
            if n_match > 0:
                st.info(f"🔗 Correlação: {n_match}/{n_total} registos com Centro identificado via Excel.")
            else:
                st.info(
                    "ℹ️ Nenhuma correspondência encontrada entre Shortname e Ação do Excel. "
                    "Carregue também o ficheiro Excel de Ações para enriquecer os dados."
                )

        if dfs_csv or dfs_acoes:
            st.session_state.quest_uploader_key += 1
            st.session_state.quest_editor_key   += 1
            st.rerun()

    st.markdown("---")

    # ── Filtros ────────────────────────────────────────────────
    df_atual = garantir_todas_colunas(st.session_state.quest_editaveis.copy())

    with st.expander("🔍 Filtrar visualização"):
        if not df_atual.empty and df_atual["Shortname"].notna().any():
            col_f1, col_f2, col_f3 = st.columns(3)

            with col_f1:
                centros       = sorted(df_atual["Centro"].dropna().unique())
                filtro_centro = st.multiselect("Centro", centros, default=centros, key="fq_centro")
            with col_f2:
                modulos       = sorted(df_atual["Módulo"].dropna().unique())
                filtro_modulo = st.multiselect("Módulo", modulos, default=modulos, key="fq_modulo")
            with col_f3:
                folhas        = sorted(df_atual["Folha"].dropna().unique())
                filtro_folha  = st.multiselect("Folha", folhas, default=folhas, key="fq_folha")

            col_f4, col_f5 = st.columns(2)
            with col_f4:
                respondentes      = sorted(df_atual["Respondente"].dropna().unique())
                filtro_respondente = st.multiselect("Respondente", respondentes, default=respondentes, key="fq_respondente")
            with col_f5:
                modalidades       = sorted(df_atual["Modalidade"].dropna().unique())
                filtro_modalidade = st.multiselect("Modalidade", modalidades, default=modalidades, key="fq_modalidade")

            col_d1, col_d2 = st.columns(2)
            df_datas       = df_atual.copy()
            df_datas["Datini_dt"] = pd.to_datetime(df_datas["Datini"], errors="coerce")
            data_min       = df_datas["Datini_dt"].min()
            data_max       = df_datas["Datini_dt"].max()

            with col_d1:
                filtro_data_ini = (
                    st.date_input("Data início (a partir de)", value=data_min.date(), key="fq_data_ini")
                    if pd.notna(data_min) else None
                )
            with col_d2:
                filtro_data_fim = (
                    st.date_input("Data início (até)", value=data_max.date(), key="fq_data_fim")
                    if pd.notna(data_max) else None
                )

            df_filtrado = df_atual.copy()
            if filtro_centro:
                df_filtrado = df_filtrado[
                    df_filtrado["Centro"].isin(filtro_centro) | df_filtrado["Centro"].isna()
                ]
            if filtro_modulo:
                df_filtrado = df_filtrado[df_filtrado["Módulo"].isin(filtro_modulo)]
            if filtro_folha:
                df_filtrado = df_filtrado[df_filtrado["Folha"].isin(filtro_folha)]
            if filtro_respondente:
                df_filtrado = df_filtrado[df_filtrado["Respondente"].isin(filtro_respondente)]
            if filtro_modalidade:
                df_filtrado = df_filtrado[df_filtrado["Modalidade"].isin(filtro_modalidade)]
            if filtro_data_ini and filtro_data_fim:
                df_datas_f = df_filtrado.copy()
                df_datas_f["Datini_dt"] = pd.to_datetime(df_datas_f["Datini"], errors="coerce")
                mask_data = df_datas_f["Datini_dt"].isna() | (
                    (df_datas_f["Datini_dt"].dt.date >= filtro_data_ini) &
                    (df_datas_f["Datini_dt"].dt.date <= filtro_data_fim)
                )
                df_filtrado = df_filtrado[mask_data.values]

            if not df_filtrado.empty:
                st.markdown(f"**📋 Visualização filtrada:** {len(df_filtrado)} registos")
                st.dataframe(
                    df_filtrado.drop(columns=["Apagar"], errors="ignore"),
                    use_container_width=True,
                    hide_index=True,
                )
            else:
                st.info("Nenhum registo corresponde aos filtros selecionados.")
        else:
            st.info("Carregue ficheiros para ativar os filtros.")

    # ── Adicionar linhas ───────────────────────────────────────
    st.markdown("---")
    st.subheader("➕ Adicionar linhas vazias")
    col_add1, col_add2 = st.columns([1, 2])
    with col_add1:
        num_linhas = st.number_input(
            "Nº de linhas", min_value=1, max_value=10000,
            value=5, step=1, key="num_linhas_quest",
        )
    with col_add2:
        if st.button("➕ Adicionar linhas", use_container_width=True):
            novas = pd.DataFrame({col: [None] * num_linhas for col in COLUNAS_DADOS})
            novas.insert(0, "Apagar", False)
            st.session_state.quest_editaveis = pd.concat(
                [st.session_state.quest_editaveis, novas], ignore_index=True
            )
            st.session_state.quest_editor_key += 1
            st.rerun()

    # ── Exportar ───────────────────────────────────────────────
    # O Excel só é regenerado quando os dados mudam (via hash).
    df_para_export = st.session_state.quest_editaveis.drop(columns=["Apagar"], errors="ignore").copy()
    tem_dados      = not df_para_export.empty and df_para_export["Shortname"].notna().any()

    if tem_dados:
        COLS_FILTRO_EXP = ["Centro", "Shortname", "Datini", "Datfim", "U_status",
                           "Respondente", "Modalidade", "Tipo"]
        COLS_REST_EXP   = [c for c in COLUNAS_DADOS if c not in COLS_FILTRO_EXP]
        ordem_exp       = [c for c in COLS_FILTRO_EXP + COLS_REST_EXP if c in df_para_export.columns]
        df_exp          = df_para_export[ordem_exp].copy()

        # ── MELHORIA 2: formatação de datas vectorizada ────────────────────────
        for col_data in ["Datini", "Datfim", "Data"]:
            if col_data in df_exp.columns:
                df_exp[col_data] = formatar_coluna_data(df_exp[col_data])

        # ── MELHORIA 3: só regenera Excel se os dados mudaram ─────────────────
        hash_atual = hash_df(df_exp)
        if st.session_state.quest_excel_hash != hash_atual:
            st.session_state.quest_excel_cache = _gerar_excel_com_filtros(df_exp)
            st.session_state.quest_excel_hash  = hash_atual

    st.markdown("---")
    st.subheader("📎 Exportar dados com filtros")
    st.caption("O ficheiro Excel inclui **filtros nativos** em todas as colunas — Centro, Ação, Data, Respondente, U_status, etc.")

    if st.session_state.quest_excel_cache is not None:
        st.download_button(
            label="📥 Descarregar Excel com filtros",
            data=st.session_state.quest_excel_cache,
            file_name="questionarios_exportados.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="download_quest_excel",
        )
    else:
        st.info("Carregue dados para activar a exportação.")

    # ── Tabela editável ────────────────────────────────────────
    st.markdown("---")
    st.subheader("✏️ Tabela de questionários")

    edited_df = st.data_editor(
        df_atual,
        use_container_width=True,
        num_rows="dynamic",
        height=420,
        key=f"quest_editor_{st.session_state.quest_editor_key}",
        column_config={
            "Apagar":      st.column_config.CheckboxColumn("Apagar", default=False),
            "Valor Médio": st.column_config.NumberColumn("Valor Médio", format="%.2f"),
        },
    )

    # ── MELHORIA 3: comparação por hash em vez de .astype(str).equals() ───────
    if hash_df(edited_df) != hash_df(df_atual):
        st.session_state.quest_editaveis  = edited_df
        st.session_state.quest_editor_key += 1
        st.rerun()

    # ── Apagar selecionados ────────────────────────────────────
    st.markdown("---")
    if st.button("✖️ Apagar selecionados", use_container_width=True):
        df   = st.session_state.quest_editaveis.copy()
        mask = df.get("Apagar", pd.Series(False, index=df.index)) == True
        if mask.any():
            df             = df[~mask].reset_index(drop=True)
            df["Apagar"]   = False
            st.session_state.quest_editaveis  = df
            st.session_state.quest_editor_key += 1
            st.rerun()
        else:
            st.warning("Nenhuma linha marcada.")

    # ── Métricas ───────────────────────────────────────────────
    df_m = st.session_state.quest_editaveis.drop(columns=["Apagar"], errors="ignore").copy()
    df_m = df_m[df_m["Shortname"].notna()]
    df_m = df_m[~df_m["Shortname"].astype(str).str.strip().isin(["", "None", "nan"])]
    df_m["Valor Médio"] = pd.to_numeric(df_m["Valor Médio"], errors="coerce")

    if not df_m.empty:
        st.markdown("---")
        st.subheader("📊 Resumo Geral")

        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("Cursos (Shortnames)", df_m["Shortname"].nunique())
        m2.metric("Centros",  df_m["Centro"].nunique()  if "Centro"  in df_m else 0)
        m3.metric("Módulos",  df_m["Módulo"].nunique()  if "Módulo"  in df_m else 0)
        m4.metric("Registos", len(df_m))
        media_geral = df_m["Valor Médio"].mean()
        m5.metric("Valor Médio Geral", f"{media_geral:.2f}" if pd.notna(media_geral) else "—")

        if all(c in df_m.columns for c in ["Módulo", "Folha", "Respondente", "Valor Médio"]):
            resumo = (
                df_m.groupby(["Módulo", "Folha", "Respondente"])["Valor Médio"]
                .mean().round(3).reset_index()
                .rename(columns={"Valor Médio": "Média"})
                .sort_values(["Módulo", "Folha"])
            )
            if not resumo.empty:
                st.subheader("📋 Médias por Módulo / Folha / Respondente")
                st.dataframe(resumo, use_container_width=True, hide_index=True)

        if "Centro" in df_m.columns and df_m["Centro"].notna().any():
            resumo_centro = (
                df_m.groupby("Centro")["Valor Médio"]
                .mean().round(3).reset_index()
                .rename(columns={"Valor Médio": "Média"})
                .sort_values("Média", ascending=False)
            )
            if not resumo_centro.empty:
                st.subheader("🏢 Médias por Centro")
                st.dataframe(resumo_centro, use_container_width=True, hide_index=True)
    else:
        st.info("ℹ️ Carregue ficheiros para ver os dados.")


if __name__ == "__main__":
    mostrar_questionarios()