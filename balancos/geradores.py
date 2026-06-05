"""
geradores.py — Geração automática do Modelo Formação Centros a partir do
Modelo Execução Física.

Colocar em: balancos/geradores.py

A lógica foi validada contra dados reais: 13/13 centros batem certo com a
contagem direta do Execução Física.
"""
import re
import io
import shutil
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent
TEMPLATE_CENTROS = SCRIPT_DIR / "Modelos" / "Modelo Formacao Centros.xlsx"
TEMPLATE_ACOES_BL = SCRIPT_DIR / "Modelos" / "Modelo Ações BL.xlsx"

# ──────────────────────────────────────────────────────────────────────────
# Lista oficial de tipologias de curso.
# 45 extraídas do preenchimento manual + MCB e MTV (confirmadas legítimas).
# SE APARECER UMA TIPOLOGIA NOVA: o gerador avisa mas NÃO a adiciona sozinho.
# Para a tornar oficial, acrescenta-a aqui à mão.
# ──────────────────────────────────────────────────────────────────────────
TIPOLOGIAS_OFICIAIS = {
    'ALM10', 'ALMA', 'APAAA', 'APAPA', 'ARD', 'ARE', 'BARM', 'BAS', 'CCPU',
    'CM', 'COPA', 'COZ', 'DDI', 'EFAEF', 'ETU', 'EUFCD', 'FETP', 'FM', 'GCF',
    'INFB', 'INFBP', 'INFBP2', 'INGP3', 'MKTD', 'SBVA', 'SPR50', 'SSPE', 'TAC',
    'TAE', 'TAFM', 'TAG', 'TAMD', 'TASG', 'TAV', 'TAVP1', 'TBP', 'TBP3',
    'TEUG', 'TEUGP', 'TPP', 'TRH', 'VIG', 'VIGA', 'VSP', 'VTVA', 'MCB', 'MTV',
}

# Colunas mínimas que o Execução Física tem de ter
COLUNAS_OBRIGATORIAS = {"U_id", "Deslocal", "Nhoras", "Total_formandos", "Desistentes"}


def _normalizar_deslocal(valor) -> str:
    """GERAL e Online/ONLINE → 'ONLINE'. Resto fica como está (sem espaços)."""
    v = str(valor).strip()
    return "ONLINE" if v.upper() in ("GERAL", "ONLINE") else v


def _derivar_tipologia(uid, avisos: set) -> str:
    """Deriva a tipologia do U_id, validando contra a lista oficial.
    Avisa (sem inventar) quando não reconhece."""
    base = str(uid).split('/')[0].replace('_BL', '').replace('_EL', '').strip().upper()
    if base in TIPOLOGIAS_OFICIAIS:
        return base
    so_letras = re.sub(r'[^A-Z]', '', base)
    if so_letras in TIPOLOGIAS_OFICIAIS:
        return so_letras
    avisos.add(f"Tipologia '{base}' (de U_id '{uid}') não está na lista oficial — incluída na mesma.")
    return base


def gerar_formacao_centros(df_execucao: pd.DataFrame, caminho_template=None):
    """
    Recebe o DataFrame do Execução Física e devolve (bytes_xlsx, avisos).
    Não grava nada em disco permanente — devolve os bytes para download.

    - bytes_xlsx: conteúdo do .xlsx pronto para st.download_button
    - avisos: lista de strings (tipologias não reconhecidas, centros sem aba)
    """
    avisos = set()
    template = Path(caminho_template) if caminho_template else TEMPLATE_CENTROS

    if not template.exists():
        raise FileNotFoundError(
            f"Template do Centros não encontrado em '{template}'. "
            f"Coloca o 'Modelo Formacao Centros.xlsx' vazio na pasta Modelos."
        )

    df = df_execucao.copy()
    df.columns = df.columns.str.strip()

    em_falta = COLUNAS_OBRIGATORIAS - set(df.columns)
    if em_falta:
        raise ValueError(
            f"O ficheiro Execução Física não tem as colunas: {sorted(em_falta)}. "
            f"Colunas encontradas: {sorted(df.columns)}"
        )

    df['centro'] = df['Deslocal'].apply(_normalizar_deslocal)
    df['tipo'] = df['U_id'].apply(lambda u: _derivar_tipologia(u, avisos))
    for c in ('Nhoras', 'Total_formandos', 'Desistentes'):
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

    # ── REGRAS DE ENCAMINHAMENTO ─────────────────────────────────────────
    # 1) _BL (b-learning) NÃO entra no Centros → vai para o documento Ações BL.
    # 2) _EL (e-learning) entra no Centros mas SEMPRE na folha ONLINE,
    #    independentemente do Deslocal de origem.
    # 3) Presenciais → folha do respetivo centro.
    uid = df['U_id'].astype(str)
    eh_bl = uid.str.contains('_BL', na=False)
    eh_el = uid.str.contains('_EL', na=False)

    n_bl = int(eh_bl.sum())
    df = df[~eh_bl].copy()
    if n_bl:
        avisos.add(f"{n_bl} ações '_BL' (b-learning) excluídas do Centros — "
                   f"destinam-se ao documento Ações BL.")

    # _EL → forçar centro para ONLINE (recalcular eh_el após remover _BL)
    eh_el = df['U_id'].astype(str).str.contains('_EL', na=False)
    n_el = int(eh_el.sum())
    df.loc[eh_el, 'centro'] = 'ONLINE'
    if n_el:
        avisos.add(f"{n_el} ações '_EL' (e-learning) colocadas na folha ONLINE.")

    # Trabalhar sobre uma cópia temporária do template
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = tmp.name
    shutil.copy(template, tmp_path)
    wb = load_workbook(tmp_path)

    abas = {a.upper(): a for a in wb.sheetnames}
    centros_sem_aba = set()

    for centro, dfc in df.groupby('centro'):
        aba = abas.get(centro.upper())
        if aba is None:
            centros_sem_aba.add(centro)
            continue
        ws = wb[aba]
        linha = 4  # dados começam na linha 4 (cabeçalho na 3)
        for tipo, dft in sorted(dfc.groupby('tipo')):
            n_acoes   = len(dft)
            horas     = int(dft['Nhoras'].sum())
            formandos = int(dft['Total_formandos'].sum())
            desist    = int(dft['Desistentes'].sum())
            volume    = horas * formandos                          # multiplicação (confirmado)
            taxa      = round(desist / formandos * 100, 2) if formandos else 0  # (desist/form)*100

            ws.cell(linha, 1, tipo)
            ws.cell(linha, 2, 0)          # Duração Média (Dias) — fica 0
            ws.cell(linha, 3, n_acoes)
            ws.cell(linha, 4, horas)
            ws.cell(linha, 5, desist)
            ws.cell(linha, 6, formandos)
            ws.cell(linha, 7, volume)
            ws.cell(linha, 8, taxa)
            linha += 1

        ultima = linha - 1
        if ultima >= 4:  # reamarrar as fórmulas do resumo (cols J/K) ao range real
            ws.cell(4, 11, f"=SUM(C4:C{ultima})")  # Nº de Ações
            ws.cell(5, 11, f"=SUM(F4:F{ultima})")  # Nº de Formandos
            ws.cell(6, 11, f"=SUM(D4:D{ultima})")  # Nº de Horas
            ws.cell(7, 11, f"=SUM(E4:E{ultima})")  # Nº de Desistências

    # Guardar para um buffer em memória (não para disco permanente)
    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)

    if centros_sem_aba:
        avisos.add(
            f"Centros presentes nos dados mas sem aba no template (ignorados): "
            f"{sorted(centros_sem_aba)}"
        )

    return buffer.getvalue(), sorted(avisos)


def _formatar_data(valor):
    """Mostra só a data (dd.mm.aaaa), descartando a hora 00:00:00."""
    if valor is None or str(valor).strip() == "":
        return ""
    try:
        return pd.to_datetime(valor, dayfirst=True).strftime("%d.%m.%Y")
    except Exception:
        # se já vier como texto com hora colada, corta no espaço
        return str(valor).split()[0]


def gerar_acoes_bl(df_execucao: pd.DataFrame, caminho_template=None):
    """
    Lista (ação a ação) as ações '_BL' do Execução Física numa folha única.
    Colunas: U_id, Datini, Datfim, Codun (sigla antes da '/'), Descun (vazio,
    para preencher à mão), U_status.

    Devolve (bytes_xlsx, avisos).
    """
    avisos = set()
    template = Path(caminho_template) if caminho_template else TEMPLATE_ACOES_BL

    if not template.exists():
        raise FileNotFoundError(
            f"Template do Ações BL não encontrado em '{template}'. "
            f"Coloca o 'Modelo Ações BL.xlsx' na pasta Modelos."
        )

    df = df_execucao.copy()
    df.columns = df.columns.str.strip()
    if "U_id" not in df.columns:
        raise ValueError("O Execução Física não tem a coluna 'U_id'.")

    df['U_id'] = df['U_id'].astype(str)
    bl = df[df['U_id'].str.contains('_BL', na=False)].copy()

    if bl.empty:
        avisos.add("Nenhuma ação '_BL' encontrada no ficheiro.")

    # Trabalhar sobre cópia do template
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = tmp.name
    shutil.copy(template, tmp_path)
    wb = load_workbook(tmp_path)
    ws = wb[wb.sheetnames[0]]   # folha única

    # Cabeçalho está na linha 1, colunas D..I (4..9):
    # D=U_id, E=Datini, F=Datfim, G=Codun, H=Descun, I=U_status
    linha = 2
    for _, r in bl.iterrows():
        uid = r['U_id']
        codun = uid.split('/')[0]               # sigla antes da '/'
        ws.cell(linha, 4, uid)
        ws.cell(linha, 5, _formatar_data(r.get('Datini')))
        ws.cell(linha, 6, _formatar_data(r.get('Datfim')))
        ws.cell(linha, 7, codun)
        ws.cell(linha, 8, "")                   # Descun — preencher à mão
        ws.cell(linha, 9, r.get('U_status', ''))
        linha += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)

    avisos.add(f"{len(bl)} ações '_BL' listadas. Coluna 'Descun' fica vazia para preenchimento manual.")
    return buffer.getvalue(), sorted(avisos)