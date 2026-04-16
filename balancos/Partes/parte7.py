# Partes/parte7.py
from pathlib import Path
import openpyxl
import re

def calcular_parte7(ano: int, regiao: str) -> dict:
    print(f"   🔍 Parte 7 - Procurando relatório para: {regiao}")

    pasta_relatorios = Path(__file__).parent.parent / "relatorios" / str(ano)
    
    if not pasta_relatorios.exists():
        print(f"   ⚠️ Pasta não encontrada: {pasta_relatorios}")
        return _valores_default_parte7(ano)  # ← PASSAR O ANO AQUI

    nome_limpo = re.sub(r'[^a-zA-Z0-9_]', '_', regiao.strip())
    possiveis_nomes = [
        f"Relatório_{nome_limpo}.xlsx",
        f"Relatorio_{nome_limpo}.xlsx",
        f"Relatório_{regiao.replace(' ', '_')}.xlsx",
        f"Relatorio_{regiao.replace(' ', '_')}.xlsx",
    ]

    caminho_excel = None
    for nome in possiveis_nomes:
        candidato = pasta_relatorios / nome
        if candidato.exists():
            caminho_excel = candidato
            print(f"   ✅ Ficheiro encontrado: {caminho_excel.name}")
            break

    if not caminho_excel:
        print(f"   ⚠️ Nenhum relatório encontrado para '{regiao}'")
        return _valores_default_parte7(ano)  # ← PASSAR O ANO AQUI

    try:
        wb = openpyxl.load_workbook(caminho_excel, data_only=True)
        if "Parte 7" not in wb.sheetnames:
            print("   ⚠️ Folha 'Parte 7' não encontrada.")
            wb.close()
            return _valores_default_parte7(ano)  # ← PASSAR O ANO AQUI

        ws = wb["Parte 7"]
        dados = {}

        # ====================== LEITURA ROBUSTA - LINHA 50 (valores) / LINHA 51 (códigos) ======================
        for col in range(1, 20):
            valor = ws.cell(row=50, column=col).value
            if not isinstance(valor, (int, float)):
                continue

            codigo = str(ws.cell(row=51, column=col).value or "").strip().upper()
            val = round(float(valor), 3)

            # ==================== MAPEAMENTO POR COLUNA + TEXTO ====================

            # Coluna 1 → Número total de respostas
            if col == 1 or any(k in codigo for k in ["QUESTIONARIOS", "IONARIOS", "RECEBIDOS"]):
                dados["NUM_TOTAL_RESPOSTAS"] = int(valor)
                dados["NUM_QUESTIONARIOS_RECEBIDOS"] = int(valor)

            # SATISFAÇÃO GLOBAL DOS FORMANDOS
            elif col == 2 or "SATISFACAO.GLOBAL.FORMANDOS" in codigo:
                dados["SATISFACAO_GLOBAL_FORMANDOS"] = val

            # SATISFAÇÃO GLOBAL DOS FORMADORES
            elif col == 3 or "SATISFACAO.GLOBAL.FORMADORES" in codigo:
                dados["SATISFACAO_GLOBAL_FORMADORES"] = val

            # SATISFAÇÃO GLOBAL DOS TUTORES
            elif col == 4 or any(k in codigo for k in ["TUTORES", "GLOBAL.TU", "GLOBAL TU", "SATISFACAO.GLOBAL.TUTORES"]):
                dados["SATISFACAO_GLOBAL_TUTORES"] = val

            # SATISFAÇÃO DA COORDENAÇÃO COM O DESEMPENHO DOS FORMADORES
            elif col == 5 or "SATISFACAO.COORDENACAO.FORMADORES" in codigo:
                dados["SATISFACAO_COORDENACAO_FORMADORES"] = val

            # SATISFAÇÃO DOS FORMANDOS COM O DESEMPENHO DOS FORMADORES
            elif col == 6 or "SATISFACAO.FORMANDOS.FORMADORES" in codigo:
                dados["SATISFACAO_FORMANDOS_FORMADORES"] = val

            # SATISFAÇÃO DOS FORMADORES COM A COORDENAÇÃO
            elif col == 7 or "SATISFACAO.FORMADORES.COORDENACAO" in codigo:
                dados["SATISFACAO_FORMADORES_COORDENACAO"] = val

            # SATISFAÇÃO DOS FORMANDOS COM A COORDENAÇÃO
            elif col == 8 or "SATISFACAO.FORMANDOS.COORDENACAO" in codigo:
                dados["SATISFACAO_FORMANDOS_COORDENACAO"] = val

            # SATISFAÇÃO DA COORDENAÇÃO (geral)
            elif col == 9 and "GESTAO" not in codigo:
                dados["SATISFACAO_COORDENACAO"] = val

            # SATISFAÇÃO DA GESTÃO
            elif col == 10 or "SATISFACAO.GESTAO" in codigo:
                dados["SATISFACAO_GESTAO"] = val

            # SATISFAÇÃO DO CLIENTE
            elif col == 11 or "SATISFACAO.CLIENTE" in codigo:
                dados["SATISFACAO_CLIENTE"] = val
            
        wb.close()

        # ====================== DEBUG ======================
        print(f"   ✅ Parte 7 lida com sucesso para {regiao}!")
        print("   📊 Valores extraídos (linha 50 / 51):")
        for k, v in sorted(dados.items()):
            print(f"      {k}: {v}")

        return {
            "NUM_TOTAL_RESPOSTAS": dados.get("NUM_TOTAL_RESPOSTAS", 0),
            "NUM_QUESTIONARIOS_RECEBIDOS": dados.get("NUM_QUESTIONARIOS_RECEBIDOS", 0),
            "SATISFACAO_GLOBAL_FORMANDOS": dados.get("SATISFACAO_GLOBAL_FORMANDOS", 0.0),
            "SATISFACAO_GLOBAL_FORMADORES": dados.get("SATISFACAO_GLOBAL_FORMADORES", 0.0),
            "SATISFACAO_GLOBAL_TUTORES": dados.get("SATISFACAO_GLOBAL_TUTORES", 0.0),
            "SATISFACAO_COORDENACAO_FORMADORES": dados.get("SATISFACAO_COORDENACAO_FORMADORES", 0.0),
            "SATISFACAO_FORMANDOS_FORMADORES": dados.get("SATISFACAO_FORMANDOS_FORMADORES", 0.0),
            "SATISFACAO_FORMADORES_COORDENACAO": dados.get("SATISFACAO_FORMADORES_COORDENACAO", 0.0),
            "SATISFACAO_FORMANDOS_COORDENACAO": dados.get("SATISFACAO_FORMANDOS_COORDENACAO", 0.0),
            "SATISFACAO_COORDENACAO": dados.get("SATISFACAO_COORDENACAO", 0.0),
            "SATISFACAO_GESTAO": dados.get("SATISFACAO_GESTAO", 0.0),
            "SATISFACAO_CLIENTE": dados.get("SATISFACAO_CLIENTE", 0.0),
            # NOTA: MEDIA_NOTAS não é definida aqui - será mantida da parte 6
            "NUM_DESISTENTES_TOTAL": 0,
            "AVALIACAO_IMPACTO_ESTAGIO": "Não aplicável",
            "SATISFACAO_FORMANDO_ESTAGIO": 0.0,
            "ANO_SEGUINTE": ano + 1,
            "VOLUME_FATURACAO_TOTAL": 0,

        }

    except Exception as e:
        print(f"   ❌ Erro ao ler Parte 7: {e}")
        return _valores_default_parte7(ano)  # ← PASSAR O ANO AQUI


def _valores_default_parte7(ano: int) -> dict:  # ← ADICIONAR PARÂMETRO ANO
    return {
        "NUM_TOTAL_RESPOSTAS": 0,
        "NUM_QUESTIONARIOS_RECEBIDOS": 0,
        "SATISFACAO_GLOBAL_FORMANDOS": 0.0,
        "SATISFACAO_GLOBAL_FORMADORES": 0.0,
        "SATISFACAO_GLOBAL_TUTORES": 0.0,
        "SATISFACAO_COORDENACAO_FORMADORES": 0.0,
        "SATISFACAO_FORMANDOS_FORMADORES": 0.0,
        "SATISFACAO_FORMADORES_COORDENACAO": 0.0,
        "SATISFACAO_FORMANDOS_COORDENACAO": 0.0,
        "SATISFACAO_COORDENACAO": 0.0,
        "SATISFACAO_GESTAO": 0.0,
        "SATISFACAO_CLIENTE": 0.0,
        # NOTA: MEDIA_NOTAS NÃO ESTÁ AQUI - será mantida da parte 6
        "NUM_DESISTENTES_TOTAL": 0,
        "AVALIACAO_IMPACTO_ESTAGIO": "Não aplicável",
        "SATISFACAO_FORMANDO_ESTAGIO": 0.0,
        "ANO_SEGUINTE": ano + 1,
        "VOLUME_FATURACAO_TOTAL": 0,

    }