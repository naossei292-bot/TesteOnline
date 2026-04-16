import pandas as pd
import openpyxl
import re
from pathlib import Path
from copy import copy
from statistics import mean

# ====================== CONFIGURAÇÕES ======================
SCRIPT_DIR = Path(__file__).parent.parent
MODELO = SCRIPT_DIR /"Modelos"/ "Modelo.xlsx"
# ALTERADO: Removido {ano} da constante - agora será construído dinamicamente na função
EXECUCAO = SCRIPT_DIR / "dados"  # Pasta base para os dados


def escrever_seguro(ws, celula, valor):
    """Escreve valor de forma segura, tratando células mescladas."""
    try:
        cell = ws[celula]
        if cell.__class__.__name__ == "MergedCell":
            for merged in list(ws.merged_cells.ranges):
                if (merged.min_row <= cell.row <= merged.max_row and 
                    merged.min_col <= cell.column <= merged.max_col):
                    ws.cell(row=merged.min_row, column=merged.min_col, value=valor)
                    return
        ws[celula] = valor
    except Exception:
        pass


def escrever_seguro_rc(ws, row, col, valor):
    """Escreve valor por (row, col) de forma segura, tratando células mescladas."""
    try:
        cell = ws.cell(row=row, column=col)
        if cell.__class__.__name__ == "MergedCell":
            for merged in list(ws.merged_cells.ranges):
                if (merged.min_row <= row <= merged.max_row and
                        merged.min_col <= col <= merged.max_col):
                    ws.cell(row=merged.min_row, column=merged.min_col, value=valor)
                    return
        ws.cell(row=row, column=col, value=valor)
    except Exception:
        pass


def extrair_codigo(texto):
    """Extrai códigos como A02, B01, A3.1.1, B3.6.1, etc."""
    if pd.isna(texto) or not isinstance(texto, str):
        return None
    texto = str(texto).upper().strip()
    match = re.search(r'([A-F]\d+(?:\.\d+)*)', texto)
    return match.group(1) if match else None


def copiar_bloco(ws, origem_inicio, origem_fim, destino_inicio):
    altura = origem_fim - origem_inicio + 1
    for i in range(altura):
        for col in range(1, 3):
            origem = ws.cell(origem_inicio + i, col)
            destino = ws.cell(destino_inicio + i, col)
            if origem.__class__.__name__ == "MergedCell":
                continue
            destino.value = origem.value
            if hasattr(origem, '_style'):
                destino._style = copy(origem._style)


def copiar_merged(ws, origem_inicio, origem_fim, destino_inicio):
    for merged in list(ws.merged_cells.ranges):
        if origem_inicio <= merged.min_row <= origem_fim:
            nova_min = destino_inicio + (merged.min_row - origem_inicio)
            nova_max = destino_inicio + (merged.max_row - origem_inicio)
            ws.merge_cells(start_row=nova_min, end_row=nova_max,
                           start_column=merged.min_col, end_column=merged.max_col)


def calcular_medias_globais_categorias(medias_por_codigo):
    categorias = {'A': [], 'B': [], 'C': [], 'D': [], 'E': [], 'F': []}
    for codigo, valor in medias_por_codigo.items():
        if codigo and len(codigo) >= 2 and valor is not None:
            letra = codigo[0].upper()
            if letra in categorias:
                categorias[letra].append(valor)
    return {letra: round(mean(valores), 3) if valores else None 
            for letra, valores in categorias.items()}


# =========================
# PREENCHER MÉDIAS NA TABELA ESQUERDA
# =========================
def preencher_medias_esquerda_por_curso(ws, mapa_medias_curso, linha_base):
    medias_categorias = calcular_medias_globais_categorias(mapa_medias_curso)
    for r in range(linha_base, linha_base + 60):
        cell_a = ws[f"A{r}"].value
        if cell_a and isinstance(cell_a, str):
            texto = str(cell_a).upper().strip()
            if "RESULTADOS POR CATEGORIA" in texto or "MÉDIA GLOBAL" in texto:
                escrever_seguro(ws, f"B{r+1}", medias_categorias.get('A'))
                escrever_seguro(ws, f"B{r+2}", medias_categorias.get('B'))
                escrever_seguro(ws, f"B{r+3}", medias_categorias.get('C'))
                escrever_seguro(ws, f"B{r+4}", medias_categorias.get('D'))
                escrever_seguro(ws, f"B{r+5}", medias_categorias.get('E'))
                print(f"   → Médias por curso preenchidas na linha {r+1} (folha {ws.title})")
                return

# =========================
# PREENCHER TABELA DA DIREITA
# =========================
def preencher_tabela_direita(ws):
    mapa_global = {}
    for r in range(1, ws.max_row + 1):
        cel_a = ws[f"A{r}"].value
        cel_b = ws[f"B{r}"].value
        codigo = extrair_codigo(cel_a)
        if codigo and isinstance(cel_b, (int, float)):
            mapa_global.setdefault(codigo, []).append(float(cel_b))

    medias_finais = {codigo: round(mean(vals), 3) 
                     for codigo, vals in mapa_global.items() if vals}

    # Tabela superior direita
    for r in range(1, ws.max_row + 1):
        cel_d = ws[f"D{r}"].value
        codigo = extrair_codigo(cel_d)
        if codigo and codigo in medias_finais:
            escrever_seguro(ws, f"E{r}", medias_finais[codigo])

    # Tabela inferior direita (A.MEDIA, B.MEDIA...)
    medias_categorias = calcular_medias_globais_categorias(medias_finais)
    for r in range(10, ws.max_row + 1):
        for col_texto in ['E', 'F']:
            cell_texto = ws[f"{col_texto}{r}"]
            if not cell_texto.value or not isinstance(cell_texto.value, str):
                continue
            texto = str(cell_texto.value).upper().strip().replace(" ", "")
            col_valor = 'F' if col_texto == 'E' else 'E'

            if any(x in texto for x in ["A.MEDIA", "AMEDIA"]):
                escrever_seguro(ws, f"{col_valor}{r}", medias_categorias.get('A'))
            elif any(x in texto for x in ["B.MEDIA", "BMEDIA"]):
                escrever_seguro(ws, f"{col_valor}{r}", medias_categorias.get('B'))
            elif any(x in texto for x in ["C.MEDIA", "CMEDIA"]):
                escrever_seguro(ws, f"{col_valor}{r}", medias_categorias.get('C'))
            elif any(x in texto for x in ["D.MEDIA", "DMEDIA"]):
                escrever_seguro(ws, f"{col_valor}{r}", medias_categorias.get('D'))
            elif any(x in texto for x in ["E.MEDIA", "EMEDIA"]):
                escrever_seguro(ws, f"{col_valor}{r}", medias_categorias.get('E'))
            elif any(x in texto for x in ["F.MEDIA", "FMEDIA", "AVALIAÇÃO GLOBAL"]):
                escrever_seguro(ws, f"{col_valor}{r}", medias_categorias.get('F'))

    print(f"   → Tabelas da direita processadas na folha: {ws.title}")

def ler_dados_folha_fonte(ws) -> dict:
    """
    Lê uma folha fonte (21a, 21b, 22a...) já preenchida e devolve um
    dicionário com os valores por placeholder.
    Agora soma todos os 'Nº de respostas' encontrados na folha.
    """
    dados = {}
    total_respostas = 0
    max_row = ws.max_row or 200

    for r in range(1, max_row + 1):
        # === Ler Nº de respostas (coluna A = 'Nº de respostas:', valor em col B) ===
        cel_a = ws.cell(r, 1).value
        if cel_a and isinstance(cel_a, str) and "nº de respostas" in str(cel_a).lower():
            val_b = ws.cell(r, 2).value
            if isinstance(val_b, (int, float)):
                total_respostas += int(round(val_b))   # soma acumulada

        # === Ler médias e indicadores (col F = placeholder, col E = valor) ===
        cel_f = ws.cell(r, 6).value
        cel_e = ws.cell(r, 5).value
        if cel_f and isinstance(cel_f, str):
            chave = str(cel_f).strip().upper().replace(" ", "")
            if isinstance(cel_e, (int, float)):
                dados[chave] = round(float(cel_e), 3)

    # Guarda o total somado
    if total_respostas > 0:
        dados["N.RESP"] = total_respostas
    else:
        dados["N.RESP"] = 0

    return dados

# =========================
# PREENCHER PARTE 7
# Lê os dados já calculados das folhas fonte e preenche a folha "Parte 7".
#
# Estrutura da Parte 7 (por blocos de linhas):
#   L4-L11  : 21a (col E = A.MEDIA..E.MEDIA, col F = N.RESP) |
#              21b (col K = A.MEDIA..F.MEDIA, col L = N.RESP)
#   L15-L22 : 22a | 22b   (mesma lógica)
#   L28-L34 : 23a | 23b
#   L40-L45 : 24a | 24b
#   L50     : linha de totais com placeholders textuais
#
# Mapeamento de colunas (base-0 → openpyxl usa base-1):
#   col 5  (E) → médias do bloco ESQUERDO  (21a, 22a, 23a, 24a)
#   col 6  (F) → N.RESP do bloco esquerdo
#   col 11 (K) → médias do bloco DIREITO   (21b, 22b, 23b, 24b)
#   col 12 (L) → N.RESP do bloco direito
# =========================
def preencher_parte7(wb: openpyxl.Workbook) -> None:
    """
    Preenche a folha 'Parte 7' com os dados já calculados nas folhas
    fonte (21a, 21b, 22a, 22b, 23a, 23b, 24a, 24b).
    Deve ser chamada DEPOIS de preencher_tabela_direita em todas as folhas.
    """
    if "Parte 7" not in wb.sheetnames:
        print("   ⚠ Folha 'Parte 7' não encontrada.")
        return

    ws7 = wb["Parte 7"]

    # ── 1. Ler dados de cada folha fonte ──────────────────────────────────────
    def dados_de(nome_folha: str) -> dict:
        """Devolve dict com os dados da folha, ou {} se não existir."""
        # Aceita nomes com e sem ponto: '21b' e '21.b'
        for variante in [nome_folha, nome_folha.replace(".", "")]:
            if variante in wb.sheetnames:
                return ler_dados_folha_fonte(wb[variante])
        print(f"   ⚠ Folha '{nome_folha}' não encontrada no workbook.")
        return {}

    d21a = dados_de("21a")
    d21b = dados_de("21b")
    d22a = dados_de("22a")
    d22b = dados_de("22b")
    d23a = dados_de("23a")
    d23b = dados_de("23b")
    d24a = dados_de("24a")
    d24b = dados_de("24b")

    print(f"   📊 Parte 7 — dados lidos:")
    print(f"      21a: {d21a}")
    print(f"      21b: {d21b}")
    print(f"      22a: {d22a}")
    print(f"      22b: {d22b}")
    print(f"      23a: {d23a}")
    print(f"      23b: {d23b}")
    print(f"      24a: {d24a}")
    print(f"      24b: {d24b}")

    # ── 2. Função auxiliar: percorre a Parte 7 e preenche onde encontrar o
    #       placeholder textual (ex: 'A.MEDIA') com o valor do dicionário certo ─
    def preencher_bloco(linha_ini: int, linha_fim: int,
                        dados_esq: dict, dados_dir: dict,
                        col_media_esq: int = 5,   # coluna E
                        col_nresp_esq: int = 6,   # coluna F
                        col_media_dir: int = 11,  # coluna K
                        col_nresp_dir: int = 12   # coluna L
                        ) -> None:
        """
        Varre as linhas [linha_ini, linha_fim] da Parte 7.
        Quando encontra um placeholder textual reconhecido numa célula,
        substitui-o pelo valor dos dicionários dados_esq / dados_dir.
        """
        PLACEHOLDERS_MEDIA = {"A.MEDIA", "B.MEDIA", "C.MEDIA",
                               "D.MEDIA", "E.MEDIA", "F.MEDIA"}

        for r in range(linha_ini, linha_fim + 1):
            # --- bloco esquerdo (cols col_media_esq e col_nresp_esq) ---
            cel_esq = ws7.cell(r, col_media_esq).value
            if cel_esq and isinstance(cel_esq, str):
                chave = str(cel_esq).strip().upper().replace(" ", "")
                if chave in PLACEHOLDERS_MEDIA and chave in dados_esq:
                    escrever_seguro_rc(ws7, r, col_media_esq, dados_esq[chave])
            cel_nresp_esq = ws7.cell(r, col_nresp_esq).value
            if cel_nresp_esq and isinstance(cel_nresp_esq, str):
                if str(cel_nresp_esq).strip().upper() == "N.RESP":
                    escrever_seguro_rc(ws7, r, col_nresp_esq, dados_esq.get("N.RESP"))

            # --- bloco direito (cols col_media_dir e col_nresp_dir) ---
            cel_dir = ws7.cell(r, col_media_dir).value
            if cel_dir and isinstance(cel_dir, str):
                chave = str(cel_dir).strip().upper().replace(" ", "")
                if chave in PLACEHOLDERS_MEDIA and chave in dados_dir:
                    escrever_seguro_rc(ws7, r, col_media_dir, dados_dir[chave])
            cel_nresp_dir = ws7.cell(r, col_nresp_dir).value
            if cel_nresp_dir and isinstance(cel_nresp_dir, str):
                if str(cel_nresp_dir).strip().upper() == "N.RESP":
                    escrever_seguro_rc(ws7, r, col_nresp_dir, dados_dir.get("N.RESP"))

    # ── 3. Preencher cada bloco de questionários ──────────────────────────────
    #   Blocos: L4-L11 → 21a/21b | L15-L22 → 22a/22b
    #           L28-L34 → 23a/23b | L40-L45 → 24a/24b
    preencher_bloco(4,  11, d21a, d21b)
    preencher_bloco(15, 22, d22a, d22b)
    preencher_bloco(28, 34, d23a, d23b)
    preencher_bloco(40, 45, d24a, d24b)

    # ── 4. Linha de totais (linha 50) ─────────────────────────────────────────
    #
    # Placeholders na linha 50 (col C em diante):
    #   C=TOTAL.RESP       → soma de todos os N.RESP
    #   D=MEDIA.EF.FORMANDOS → média global formandos (21a+21b+22a+22b)
    #   E=MEDIA.D.23A      → D.MEDIA da folha 23a
    #   F=MEDIA.D.23B      → D.MEDIA da folha 23b
    #   G=MEDIA.D.24       → D.MEDIA da folha 24a (e 24b se existir)
    #   H=MEDIA.D.FORMANDOS → D.MEDIA formandos (21a+21b D.MEDIA)
    #   I=MEDIA.B.23       → B.MEDIA da folha 23a (ou média 23a+23b)
    #   J=MEDIA.B.FORMANDOS → B.MEDIA formandos (21a+21b+22a+22b)
    #   K=MEDIA.B.21.22.23 → B.MEDIA de 21+22+23
    #   L=MEDIA.B.21.22.23 → idem (duplicado no modelo)
    #   M=MEDIA.EF.FORMANDOS → repetição da média global formandos

    def media_segura(*valores):
        """Média de valores não-None."""
        vals = [v for v in valores if isinstance(v, (int, float))]
        return round(mean(vals), 3) if vals else None

    def soma_segura(*valores):
        """Soma de valores não-None."""
        vals = [v for v in valores if isinstance(v, (int, float))]
        return sum(vals) if vals else None

    total_resp = soma_segura(
        d21a.get("N.RESP"), d21b.get("N.RESP"),
        d22a.get("N.RESP"), d22b.get("N.RESP"),
        d23a.get("N.RESP"), d23b.get("N.RESP"),
        d24a.get("N.RESP"), d24b.get("N.RESP"),
    )

    # Média global formandos = média das E.MEDIA de 21a, 21b, 22a, 22b
    media_ef_formandos = media_segura(
        d21a.get("E.MEDIA"), d21b.get("F.MEDIA"),   # 21b usa F.MEDIA como avaliação global
        d22a.get("E.MEDIA"), d22b.get("F.MEDIA"),
    )

    media_d_23a       = d23a.get("D.MEDIA")
    media_d_23b       = d23b.get("E.MEDIA")         # 23b: D é Plataforma Moodle, E é Avaliação global
    media_d_24        = media_segura(d24a.get("D.MEDIA"), d24b.get("D.MEDIA"))
    media_d_formandos = media_segura(d21a.get("D.MEDIA"), d21b.get("D.MEDIA"))
    media_b_23        = media_segura(d23a.get("B.MEDIA"), d23b.get("B.MEDIA"))
    media_b_formandos = media_segura(
        d21a.get("B.MEDIA"), d21b.get("B.MEDIA"),
        d22a.get("B.MEDIA"), d22b.get("B.MEDIA"),
    )
    media_b_21_22_23  = media_segura(
        d21a.get("B.MEDIA"), d21b.get("B.MEDIA"),
        d22a.get("B.MEDIA"), d22b.get("B.MEDIA"),
        d23a.get("B.MEDIA"), d23b.get("B.MEDIA"),
    )

    # Mapa: placeholder_texto → (coluna_openpyxl, valor)
    # Colunas: C=3, D=4, E=5, F=6, G=7, H=8, I=9, J=10, K=11, L=12, M=13
    mapa_totais = {
        "TOTAL.RESP":         (3,  total_resp),
        "MEDIA.EF.FORMANDOS": (4,  media_ef_formandos),
        "MEDIA.D.23A":        (5,  media_d_23a),
        "MEDIA.D.23B":        (6,  media_d_23b),
        "MEDIA.D.24":         (7,  media_d_24),
        "MEDIA.D.FORMANDOS":  (8,  media_d_formandos),
        "MEDIA.B.23":         (9,  media_b_23),
        "MEDIA.B.FORMANDOS":  (10, media_b_formandos),
        "MEDIA.B.21.22.23":   (11, media_b_21_22_23),
        # col 12 é duplicado de MEDIA.B.21.22.23 no modelo — preenchemos também
        "MEDIA.B.21.22.23_L": (12, media_b_21_22_23),
        # col 13 repete MEDIA.EF.FORMANDOS
        "MEDIA.EF.FORMANDOS_M": (13, media_ef_formandos),
    }

    # Percorrer linha 50 e substituir placeholders pelos valores calculados
    linha_totais = 50
    for r in range(linha_totais - 2, linha_totais + 3):  # tolerância de ±2 linhas
        for col in range(3, 14):
            cel = ws7.cell(r, col).value
            if cel and isinstance(cel, str):
                chave = str(cel).strip().upper().replace(".", "_").replace(" ", "")
                # Mapear variante sem sufixo _L/_M
                for placeholder, (coluna, valor) in mapa_totais.items():
                    chave_clean = placeholder.replace(".", "_").replace(" ", "")
                    if chave_clean in chave or chave in chave_clean:
                        if valor is not None:
                            escrever_seguro_rc(ws7, r, col, valor)
                        break

    print(f"   ✅ Parte 7 preenchida com sucesso.")
    print(f"      Total respostas: {total_resp}")
    print(f"      Média global formandos: {media_ef_formandos}")

# =========================
# FUNÇÃO PRINCIPAL (ALTERADA)
# =========================
def preparar_dados_moodle(ano):
    # ALTERADO: Construir caminhos dinâmicos com o ano
    pasta_dados_ano = SCRIPT_DIR / f"dados/{ano}"
    caminho_moodle = pasta_dados_ano / f"moodle {ano}.xlsx"
    caminho_execucao = pasta_dados_ano / "Modelo Execução Fisica.xlsx"
    
    # ALTERADO: Verificar se os ficheiros existem
    if not caminho_moodle.exists():
        raise FileNotFoundError(f"Ficheiro não encontrado: {caminho_moodle}")
    if not caminho_execucao.exists():
        raise FileNotFoundError(f"Ficheiro não encontrado: {caminho_execucao}")
    
    print(f"📁 A processar ano {ano}")
    print(f"   - Moodle: {caminho_moodle}")
    print(f"   - Execução: {caminho_execucao}")
    
    df_moodle = pd.read_excel(str(caminho_moodle))
    df_exec = pd.read_excel(str(caminho_execucao))  # ALTERADO: usar caminho dinâmico

    df_moodle.columns = df_moodle.columns.str.strip()
    df_exec.columns = df_exec.columns.str.strip()

    colunas = {c.lower(): c for c in df_exec.columns}
    u_id_col = colunas.get("u_id") or colunas.get("u_id ")

    if not u_id_col:
        raise Exception("Coluna U_id não encontrada no ficheiro de execução")

    df_moodle["shortname"] = df_moodle["shortname"].astype(str).str.strip()
    df_exec[u_id_col] = df_exec[u_id_col].astype(str).str.strip()

    df = df_moodle.merge(
        df_exec[[u_id_col, "Total_formandos", "Deslocal"]],
        left_on="shortname", right_on=u_id_col, how="left"
    )

    pasta_ano = SCRIPT_DIR / "relatorios" / str(ano)
    pasta_ano.mkdir(parents=True, exist_ok=True)

    for local in df["Deslocal"].dropna().unique():
        print(f"\n📍 A criar relatório: {local}")

        wb = openpyxl.load_workbook(str(MODELO))
        dados_local = df[df["Deslocal"] == local]

        todas_medias_subcategorias = {}

        for folha in wb.sheetnames:
            if folha in ["Resumo", "Folha1", "Parte 7"]:
                continue

            ws = wb[folha]
            modulo = folha.lower().strip()

            dados_modulo = dados_local[
                dados_local["modulo"].astype(str).str.lower().str.contains(modulo, na=False)
            ]

            if dados_modulo.empty:
                continue

            cursos = dados_modulo["shortname"].unique()

            for i, curso in enumerate(cursos):
                linha_base = 1 + (i * 36)

                if i > 0:
                    ws.insert_rows(linha_base, 36)
                    copiar_bloco(ws, 1, 36, linha_base)
                    copiar_merged(ws, 1, 36, linha_base)

                dados_curso = dados_modulo[dados_modulo["shortname"] == curso]

                # Cabeçalho
                total_formandos = dados_curso["Total_formandos"].dropna().iloc[0] if not dados_curso["Total_formandos"].dropna().empty else 0
                respostas = int(round(dados_curso["contador"].mean(), 0)) if not dados_curso.empty else 0
                percentagem = round(respostas / total_formandos, 2) if total_formandos > 0 else 0

                ws[f"B{linha_base+1}"] = curso
                ws[f"B{linha_base+2}"] = f"REF_{curso}"
                ws[f"B{linha_base+3}"] = total_formandos
                ws[f"B{linha_base+4}"] = respostas
                ws[f"B{linha_base+5}"] = percentagem
                if ws[f"B{linha_base+5}"].value is not None:
                    ws[f"B{linha_base+5}"].number_format = '0%'

                # Mapa de médias
                mapa_medias = {}
                for _, row in dados_curso.iterrows():
                    codigo = extrair_codigo(row.get("nitem"))
                    if codigo:
                        mapa_medias.setdefault(codigo, []).append(row["media"])

                for codigo in mapa_medias:
                    mapa_medias[codigo] = round(mean(mapa_medias[codigo]), 2)

                # Preencher coluna B
                for linha in range(linha_base, linha_base + 40):
                    texto = ws[f"A{linha}"].value
                    codigo = extrair_codigo(texto)
                    if codigo and codigo in mapa_medias:
                        ws[f"B{linha}"] = mapa_medias[codigo]

                preencher_medias_esquerda_por_curso(ws, mapa_medias, linha_base)

                # Acumular para Resumo
                for codigo, valor in mapa_medias.items():
                    todas_medias_subcategorias.setdefault(codigo, []).append(valor)

            preencher_tabela_direita(ws)

        # ====================== RESUMO GLOBAL ======================
        medias_finais_sub = {codigo: round(mean(lista), 3) 
                             for codigo, lista in todas_medias_subcategorias.items() if lista}
        medias_categorias = calcular_medias_globais_categorias(medias_finais_sub)

        if "Resumo" in wb.sheetnames:
            ws_resumo = wb["Resumo"]
            escrever_seguro(ws_resumo, "B26", medias_categorias.get('A'))
            escrever_seguro(ws_resumo, "B27", medias_categorias.get('B'))
            escrever_seguro(ws_resumo, "B28", medias_categorias.get('C'))
            escrever_seguro(ws_resumo, "B29", medias_categorias.get('D'))
            escrever_seguro(ws_resumo, "B30", medias_categorias.get('E'))

        # ====================== PARTE 7 ======================
        # Chamada DEPOIS de preencher_tabela_direita em todas as folhas,
        # para que os valores já estejam calculados nas folhas fonte.
        print(f"\n   🔄 A preencher Parte 7...")
        preencher_parte7(wb)

        # ====================== GUARDAR FICHEIRO ======================
        nome_base = str(local).strip()
        nome_limpo = re.sub(r'[^a-zA-Z0-9\s\-]', '', nome_base)
        nome_limpo = re.sub(r'\s+', '_', nome_limpo)
        nome_limpo = re.sub(r'_+', '_', nome_limpo)
        nome_limpo = nome_limpo.strip('_')

        if len(nome_limpo) > 70:
            nome_limpo = nome_limpo[:70]

        caminho = pasta_ano / f"Relatório_{nome_limpo}.xlsx"
        
        print(f"   → A guardar como: Relatório_{nome_limpo}.xlsx")

        try:
            wb.save(str(caminho))
            print(f"✅ Criado com sucesso: {caminho}")
        except PermissionError:
            print(f"❌ Erro: Ficheiro aberto ou bloqueado: {caminho}")
        except Exception as e:
            print(f"❌ Erro ao guardar {caminho}: {e}")
        finally:
            wb.close()

    print("\n🚀 Todos os relatórios criados com sucesso!")